import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

import export_plaxis_data as core


class ExcelSheetSplittingTests(unittest.TestCase):
    def test_multisheet_writer_splits_oversized_dataframes(self):
        df = pd.DataFrame(
            {
                "Step": list(range(12)),
                "Value": [float(i) for i in range(12)],
            }
        )
        messages = []

        with tempfile.TemporaryDirectory() as tmp_dir:
            out_path = Path(tmp_dir) / "large.xlsx"
            with patch.object(core, "EXCEL_DATA_ROWS_PER_SHEET", 5), patch.object(
                core,
                "EXCEL_MAX_ROWS",
                6,
            ):
                core._write_multisheet_workbook(
                    out_path,
                    [("NodeTimeHistoryLong", df)],
                    logger=messages.append,
                )

            wb = load_workbook(out_path, read_only=True)
            try:
                self.assertEqual(
                    wb.sheetnames,
                    [
                        "NodeTimeHistoryLong_01",
                        "NodeTimeHistoryLong_02",
                        "NodeTimeHistoryLong_03",
                    ],
                )
                self.assertEqual(wb["NodeTimeHistoryLong_01"].max_row, 6)
                self.assertEqual(wb["NodeTimeHistoryLong_02"].max_row, 6)
                self.assertEqual(wb["NodeTimeHistoryLong_03"].max_row, 3)
            finally:
                wb.close()

        self.assertTrue(any("splitting into 3 sheets" in message for message in messages))


if __name__ == "__main__":
    unittest.main()
