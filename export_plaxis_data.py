import argparse
import math
import re
import time
from pathlib import Path

import numpy as np
import pandas as pd


def parse_args():
    parser = argparse.ArgumentParser(
        description="Export PLAXIS spectrum table or node time history to file."
    )
    subparsers = parser.add_subparsers(dest="mode", required=True)

    p_spectrum = subparsers.add_parser(
        "spectrum-gui",
        help="Read the active Points table via GUI Copy button and export to Excel.",
    )
    p_spectrum.add_argument(
        "--out",
        default=r"C:\Users\PC\OneDrive\Desktop\PLAXIS_spectrum_from_points.xlsx",
        help="Output .xlsx path.",
    )
    p_spectrum.add_argument(
        "--output-window-title",
        default=r".*PLAXIS 2D Ultimate Output.*",
        help="Regex for main Output window title.",
    )
    p_spectrum.add_argument(
        "--points-window-title",
        default=r".*(Points|Curve points|Curves manager|Select points).*",
        help="Regex for points/curves window title.",
    )
    p_spectrum.add_argument(
        "--combo-index",
        type=int,
        default=0,
        help="If multiple combobox controls exist in Points window, pick this index.",
    )
    p_spectrum.add_argument(
        "--copy-button",
        default="Copy",
        help='Button title used for table copy (default: "Copy").',
    )
    p_spectrum.add_argument(
        "--x-col",
        default="Period_s",
        help="X-axis output column name (default: Period_s).",
    )
    p_spectrum.add_argument(
        "--y-col",
        default="PSA_g",
        help="Y-axis output column name (default: PSA_g).",
    )
    p_spectrum.add_argument(
        "--wait",
        type=float,
        default=0.25,
        help="Delay between GUI actions in seconds.",
    )
    p_spectrum.add_argument(
        "--node",
        action="append",
        default=[],
        help="Node name from Points combobox. Can be repeated.",
    )

    p_history = subparsers.add_parser(
        "timehistory-api",
        help="Read node time history from Output scripting server.",
    )
    p_history.add_argument("--host", default="localhost")
    p_history.add_argument("--port", type=int, default=10001)
    p_history.add_argument("--password", required=True)
    p_history.add_argument("--phase-index", type=int, default=-1)
    p_history.add_argument(
        "--phase-name",
        default="",
        help="Phase identification name, for example: Phase_6",
    )
    p_history.add_argument(
        "--result-type",
        default="Soil.Ax",
        help="ResultTypes path, for example: Soil.Ax, Soil.Ux, Soil.Uy",
    )
    p_history.add_argument(
        "--point",
        action="append",
        default=[],
        help='Target coordinate in "x,y" format. Can be repeated.',
    )
    p_history.add_argument(
        "--points-file",
        default="",
        help="CSV/XLSX containing x,y columns. Used with or instead of --point.",
    )
    p_history.add_argument(
        "--all-nodes",
        action="store_true",
        help="Export all nodes (can be very large).",
    )
    p_history.add_argument(
        "--out",
        default=r"C:\Users\PC\OneDrive\Desktop\PLAXIS_timehistory.xlsx",
        help="Output .xlsx or .csv path.",
    )
    p_history.add_argument(
        "--time-col",
        default="DynamicTime",
        help="Time column name in output table.",
    )

    p_curve = subparsers.add_parser(
        "curvepoints-api",
        help="Export CurvePoints time history and computed spectrum from Output API.",
    )
    p_curve.add_argument("--host", default="localhost")
    p_curve.add_argument("--port", type=int, default=10001)
    p_curve.add_argument("--password", required=True)
    p_curve.add_argument("--phase-index", type=int, default=-1)
    p_curve.add_argument("--phase-name", default="", help="Phase identification name.")
    p_curve.add_argument(
        "--result-type",
        default="Soil.Ax",
        help="Acceleration ResultTypes path for time series input, for example Soil.Ax.",
    )
    p_curve.add_argument(
        "--velocity-result-type",
        default="",
        help=(
            "Velocity ResultTypes path (for example Soil.Vx). "
            "Empty=auto from acceleration type, none/off/- to disable."
        ),
    )
    p_curve.add_argument(
        "--displacement-result-type",
        default="",
        help=(
            "Displacement ResultTypes path (for example Soil.Ux). "
            "Empty=auto from acceleration type, none/off/- to disable."
        ),
    )
    p_curve.add_argument(
        "--curvepoint-id",
        action="append",
        default=[],
        help="Curve point id from list_curve_points_api. Can be repeated. If empty, all are used.",
    )
    p_curve.add_argument(
        "--out",
        default=r"C:\Users\PC\OneDrive\Desktop\PLAXIS_curvepoints_time_spectrum.xlsx",
        help="Output .xlsx path.",
    )
    p_curve.add_argument("--time-col", default="DynamicTime")
    p_curve.add_argument("--damping", type=float, default=0.05, help="Damping ratio (0-1).")
    p_curve.add_argument("--period-start", type=float, default=0.01)
    p_curve.add_argument("--period-end", type=float, default=3.00)
    p_curve.add_argument("--period-step", type=float, default=0.01)
    p_curve.add_argument(
        "--split-output-files",
        action="store_true",
        help="Write TimeHistory/Spectrum/NodeMap/_Status to separate files.",
    )

    return parser.parse_args()


def as_float(value):
    try:
        return float(value.value)
    except Exception:
        return float(value)


def xy_of(obj):
    for ax in ("x", "X"):
        for ay in ("y", "Y"):
            if hasattr(obj, ax) and hasattr(obj, ay):
                return as_float(getattr(obj, ax)), as_float(getattr(obj, ay))
    raise RuntimeError("Node object has no x/y attributes.")


def sanitize_sheet_name(name):
    clean = re.sub(r'[\\/*?:\[\]]', "_", str(name)).strip()
    return (clean[:31] or "Node")


def safe_label(name):
    return re.sub(r"[^0-9a-zA-Z_]+", "_", name).strip("_")


def parse_points_table(text):
    rows = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        parts = line.split()
        if len(parts) < 2:
            continue
        try:
            if len(parts) >= 3 and re.fullmatch(r"[+-]?\d+", parts[0]):
                idx = int(parts[0])
                x = float(parts[1].replace(",", "."))
                y = float(parts[2].replace(",", "."))
            else:
                idx = len(rows)
                x = float(parts[0].replace(",", "."))
                y = float(parts[1].replace(",", "."))
            rows.append((idx, x, y))
        except ValueError:
            continue
    if not rows:
        raise RuntimeError(
            "No numeric rows parsed from copied table. "
            "Keep Points table active and validate one manual Copy."
        )
    return pd.DataFrame(rows, columns=["Index", "X", "Y"])


def _error_text(exc):
    text = str(exc).strip()
    return text if text else repr(exc)


def _open_output_server(host, port, password):
    from plxscripting.easy import new_server

    last_exc = None
    for attempt in (1, 2):
        try:
            return new_server(host, port, password=password)
        except Exception as exc:
            last_exc = exc
            err_text = _error_text(exc)
            # Known intermittent PLAXIS session handshake issue; retry once.
            if "Reply code is different from what was sent" in err_text and attempt == 1:
                time.sleep(0.4)
                continue
            raise
    raise last_exc


def _safe_close_server(server):
    # In plxscripting, server.close() closes the active PLAXIS project.
    # Keep the project/session open and let Python GC handle the proxy object.
    _ = server


def _versioned_output_path(path, version):
    return path.with_name(f"{path.stem}_v{version}{path.suffix}")


def _open_excel_writer_with_fallback(out_path, logger=None, max_versions=30):
    candidates = [out_path]
    for i in range(2, max_versions + 2):
        candidates.append(_versioned_output_path(out_path, i))

    for idx, candidate in enumerate(candidates):
        try:
            writer = pd.ExcelWriter(candidate, engine="openpyxl")
            if idx > 0 and logger is not None:
                logger(f"Output file locked, writing to {candidate}")
            return writer, candidate
        except PermissionError:
            continue

    raise RuntimeError(
        f"Cannot write output file: {out_path}. "
        "Close the file in Excel/OneDrive and retry."
    )


def _save_csv_with_fallback(df, out_path, logger=None, max_versions=30):
    candidates = [out_path]
    for i in range(2, max_versions + 2):
        candidates.append(_versioned_output_path(out_path, i))

    for idx, candidate in enumerate(candidates):
        try:
            df.to_csv(candidate, index=False)
            if idx > 0 and logger is not None:
                logger(f"Output file locked, writing to {candidate}")
            return candidate
        except PermissionError:
            continue

    raise RuntimeError(
        f"Cannot write output file: {out_path}. "
        "Close the file in Excel/OneDrive and retry."
    )


def _write_single_sheet_xlsx_with_fallback(df, out_path, sheet_name, logger=None):
    writer, out_path_final = _open_excel_writer_with_fallback(out_path, logger=logger)
    with writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return out_path_final


def _ensure_active_output_project(g_o):
    try:
        _ = list(g_o.Phases)
    except Exception as exc:
        err = _error_text(exc)
        if "No active project" in err:
            raise RuntimeError(
                "PLAXIS Output has no active project. "
                "Open the calculated project in Output and try again."
            ) from exc
        raise


def _ensure_output_result_types(g_o):
    try:
        _ = g_o.ResultTypes
    except Exception as exc:
        err = _error_text(exc)
        if "ResultTypes" in err:
            raise RuntimeError(
                "Connected port is not PLAXIS Output API (ResultTypes missing). "
                "Use the Output remote scripting port."
            ) from exc
        raise


def _unique_sheet_name(base_name, used_names):
    candidate = sanitize_sheet_name(base_name)
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate
    for idx in range(2, 1000):
        suffix = f"_{idx}"
        trimmed = candidate[: max(1, 31 - len(suffix))]
        alt = f"{trimmed}{suffix}"
        if alt not in used_names:
            used_names.add(alt)
            return alt
    raise RuntimeError(f"Could not allocate unique sheet name for '{base_name}'.")


def click_button(window, title):
    btn = window.child_window(title=title, control_type="Button")
    btn.wait("exists enabled visible", timeout=5)
    btn.click_input()


def _match_node_name(text):
    if not text:
        return None
    m = re.search(r"\bNode\s+\d+\b", str(text), flags=re.IGNORECASE)
    return m.group(0) if m else None


def _node_number(text):
    m = re.search(r"\bNode\s+(\d+)\b", str(text), flags=re.IGNORECASE)
    return m.group(1) if m else None


def _unique_keep_order(items):
    seen = set()
    out = []
    for item in items:
        key = str(item).strip()
        if not key:
            continue
        if key in seen:
            continue
        seen.add(key)
        out.append(key)
    return out


def _window_titles_snapshot(desktop, limit=25):
    titles = []
    for win in desktop.windows():
        try:
            title = win.window_text().strip()
        except Exception:
            continue
        if title:
            titles.append(title)
    return _unique_keep_order(titles)[:limit]


def _resolve_points_window(desktop, points_window_title, output_window=None):
    patterns = [
        points_window_title,
        r".*Curves manager.*",
        r".*Curve points.*",
        r".*Select points.*",
        r".*Points.*",
    ]
    for pattern in _unique_keep_order(patterns):
        try:
            win = desktop.window(title_re=pattern)
            win.wait("visible", timeout=3)
            return win
        except Exception:
            continue
    if output_window is not None:
        return output_window
    titles = _window_titles_snapshot(desktop)
    joined = "; ".join(titles) if titles else "(no visible windows)"
    raise RuntimeError(
        "Could not find points window. "
        f"Tried title regex: {points_window_title}. "
        f"Visible windows: {joined}"
    )


def _nodes_from_combobox(points_window, combo_index, wait_sec):
    combos = []
    try:
        combos.append(points_window.child_window(control_type="ComboBox", found_index=combo_index))
    except Exception:
        pass
    try:
        combos.extend(points_window.descendants(control_type="ComboBox"))
    except Exception:
        pass

    names = []
    used_handles = set()
    for combo in combos:
        handle = getattr(getattr(combo, "element_info", None), "handle", None)
        if handle in used_handles:
            continue
        used_handles.add(handle)
        try:
            combo.wait("exists enabled", timeout=2)
            combo.expand()
            time.sleep(max(0.1, wait_sec))
            items = combo.descendants(control_type="ListItem")
            raw_names = [it.window_text().strip() for it in items if it.window_text().strip()]
            node_names = []
            for raw_name in raw_names:
                node_name = _match_node_name(raw_name)
                if node_name:
                    node_names.append(node_name)
            # Strict mode: ignore comboboxes that do not contain node labels
            # (for example top phase/step combobox).
            names.extend(node_names)
            combo.collapse()
        except Exception:
            continue
    return _unique_keep_order(names)


def _nodes_from_descendants(points_window):
    names = []
    try:
        controls = points_window.descendants()
    except Exception:
        controls = []
    for control in controls:
        try:
            text = control.window_text().strip()
        except Exception:
            continue
        node_name = _match_node_name(text)
        if node_name:
            names.append(node_name)
    return _unique_keep_order(names)


def _get_points_combobox(points_window, combo_index):
    candidates = []
    try:
        candidates.append(points_window.child_window(control_type="ComboBox", found_index=combo_index))
    except Exception:
        pass
    try:
        candidates.extend(points_window.descendants(control_type="ComboBox"))
    except Exception:
        pass

    unique = []
    used_handles = set()
    for combo in candidates:
        handle = getattr(getattr(combo, "element_info", None), "handle", None)
        if handle in used_handles:
            continue
        used_handles.add(handle)
        unique.append(combo)

    for combo in unique:
        try:
            combo.wait("visible enabled", timeout=2)
            combo.expand()
            time.sleep(0.15)
            items = combo.descendants(control_type="ListItem")
            combo.collapse()
            if any(_match_node_name(it.window_text()) for it in items):
                return combo
        except Exception:
            continue
    raise RuntimeError(
        "Node combobox not found. "
        "The visible comboboxes are likely phase/step selectors. "
        "Open a points window/control that contains node names (Node ####)."
    )


def _get_node_tabitems(points_window):
    tabs_by_num = {}
    try:
        tab_items = points_window.descendants(control_type="TabItem")
    except Exception:
        tab_items = []
    for tab in tab_items:
        try:
            text = tab.window_text().strip()
        except Exception:
            continue
        num = _node_number(text)
        if not num:
            continue
        if num not in tabs_by_num:
            tabs_by_num[num] = tab
    return tabs_by_num


def _build_node_selector(points_window, combo_index):
    # Fast path 1: classic node combobox.
    try:
        combo = _get_points_combobox(points_window, combo_index)

        def select_from_combo(node_name):
            target_num = _node_number(node_name)
            try:
                combo.select(node_name)
                return "combobox"
            except Exception:
                if not target_num:
                    raise
                combo.expand()
                time.sleep(0.12)
                items = combo.descendants(control_type="ListItem")
                for item in items:
                    text = item.window_text().strip()
                    if _node_number(text) == target_num:
                        try:
                            item.click_input()
                        except Exception:
                            combo.select(text)
                        combo.collapse()
                        return "combobox:number-match"
                combo.collapse()
                raise RuntimeError(f"Node '{node_name}' not found in combobox.")

        return "combobox", select_from_combo
    except Exception:
        pass

    # Fast path 2: Points window tabs (Node #### *).
    tabs_by_num = _get_node_tabitems(points_window)
    if tabs_by_num:

        def select_from_tab(node_name):
            target_num = _node_number(node_name)
            if not target_num or target_num not in tabs_by_num:
                raise RuntimeError(f"Node '{node_name}' not found in Points tabs.")
            points_window.set_focus()
            tabs_by_num[target_num].click_input()
            return "tab"

        return "tab", select_from_tab

    # Fallback: generic slow search.
    return "fallback", lambda node_name: _select_node_in_points_window(
        points_window, node_name, combo_index
    )


def _select_node_in_points_window(points_window, node_name, combo_index):
    target_num = _node_number(node_name)

    # Preferred path: select node from the node combobox (classic Points dialog).
    try:
        combo = _get_points_combobox(points_window, combo_index)
        try:
            combo.select(node_name)
            return "combobox"
        except Exception:
            if target_num:
                combo.expand()
                time.sleep(0.15)
                items = combo.descendants(control_type="ListItem")
                for item in items:
                    text = item.window_text().strip()
                    if _node_number(text) == target_num:
                        try:
                            item.click_input()
                        except Exception:
                            combo.select(text)
                        combo.collapse()
                        return "combobox:number-match"
                combo.collapse()
    except Exception:
        pass

    # Fallback path: select node tab/row from Curves manager or Points window.
    for control_type in ("TabItem", "DataItem", "ListItem", "Text"):
        try:
            if target_num:
                item = points_window.child_window(
                    title_re=rf".*\bNode\s+{target_num}\b.*", control_type=control_type
                )
            else:
                item = points_window.child_window(title=node_name, control_type=control_type)
            item.wait("exists visible enabled", timeout=1.5)
            item.click_input()
            return f"points:{control_type}"
        except Exception:
            continue

    # Last attempt: scan all descendants and match node number/text.
    try:
        for item in points_window.descendants():
            try:
                text = item.window_text().strip()
            except Exception:
                continue
            matched = False
            if target_num and _node_number(text) == target_num:
                matched = True
            elif text == node_name:
                matched = True
            if matched:
                try:
                    item.click_input()
                    return "points:descendant"
                except Exception:
                    continue
    except Exception:
        pass

    raise RuntimeError(f"Could not select node '{node_name}' in points window.")


def _copy_current_curve_data(points_window, output_window, copy_button, wait_sec):
    # Try explicit Copy button in points/manager dialogs first.
    for win in (points_window, output_window):
        try:
            click_button(win, copy_button)
            time.sleep(wait_sec)
            return "button"
        except Exception:
            continue

    # Fallback: focus Output and send Ctrl+C.
    try:
        from pywinauto.keyboard import send_keys

        output_window.set_focus()
        send_keys("^c")
        time.sleep(wait_sec)
        return "ctrl+c"
    except Exception as exc:
        raise RuntimeError("Could not trigger copy action (Copy button/Ctrl+C failed).") from exc


def _focus_points_table(points_window):
    # Prefer explicit grid controls if present.
    for control_type in ("DataGrid", "Table"):
        try:
            grids = points_window.descendants(control_type=control_type)
        except Exception:
            grids = []
        for grid in grids:
            try:
                grid.click_input()
                return True
            except Exception:
                continue

    # Fallback: click first visible numeric data cell.
    try:
        items = points_window.descendants(control_type="DataItem")
    except Exception:
        items = []
    for item in items:
        try:
            text = item.window_text().strip()
        except Exception:
            continue
        if re.fullmatch(r"[+-]?\d+(?:[.,]\d+)?", text):
            try:
                item.click_input()
                return True
            except Exception:
                continue
    return False


def _select_all_points_table(points_window, output_window, wait_sec):
    from pywinauto.keyboard import send_keys

    for win in (points_window, output_window):
        try:
            win.set_focus()
        except Exception:
            continue
        _focus_points_table(points_window)
        try:
            send_keys("^a")
            time.sleep(max(0.05, wait_sec * 0.5))
            return True
        except Exception:
            continue
    return False


def _wait_for_clipboard_update(previous_text, timeout_sec=1.5):
    import pyperclip

    end_time = time.time() + timeout_sec
    last_text = previous_text
    while time.time() < end_time:
        try:
            current = pyperclip.paste()
        except Exception:
            current = last_text
        if current and current != previous_text:
            return current
        last_text = current
        time.sleep(0.03)
    return last_text


def _collect_points_window_nodes(points_window, combo_index, wait_sec):
    names = _nodes_from_combobox(points_window, combo_index, wait_sec)
    if names:
        return names
    return _nodes_from_descendants(points_window)


def list_points_nodes(output_window_title, points_window_title, combo_index, wait_sec=0.2):
    from pywinauto import Desktop

    desktop = Desktop(backend="uia")
    output_window = desktop.window(title_re=output_window_title)
    output_window.wait("visible", timeout=10)

    points_window = _resolve_points_window(desktop, points_window_title, output_window=output_window)
    node_names = _collect_points_window_nodes(points_window, combo_index, wait_sec)
    if not node_names:
        raise RuntimeError(
            "No node names found in points window. "
            "Open Curves manager/Points window and keep node list visible."
        )
    return node_names


def run_spectrum_gui(args, logger=print):
    import pyperclip
    from pywinauto import Desktop

    desktop = Desktop(backend="uia")
    output_window = desktop.window(title_re=args.output_window_title)
    output_window.wait("visible", timeout=10)

    points_window = _resolve_points_window(
        desktop, args.points_window_title, output_window=output_window
    )
    all_nodes = _collect_points_window_nodes(points_window, args.combo_index, args.wait)

    if args.node:
        requested = [name.strip() for name in args.node if name.strip()]
        node_names = [name for name in all_nodes if name in requested]
        if not node_names:
            raise RuntimeError("Requested nodes were not found in Points combobox.")
    else:
        node_names = all_nodes

    if not node_names:
        raise RuntimeError("No node entries found in Points combobox.")

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    selector_mode, select_node = _build_node_selector(points_window, args.combo_index)
    logger(f"Node selector mode: {selector_mode}")

    merged = None
    used_sheet_names = set()
    status_rows = []
    success_count = 0
    writer, out_path_final = _open_excel_writer_with_fallback(out_path, logger=logger)
    with writer:
        for node_name in node_names:
            select_mode = ""
            copy_mode = ""
            row_count = 0
            try:
                select_mode = select_node(node_name)
                time.sleep(args.wait)

                # Points window can keep an old/partial selection after tab switch.
                # Force select-all before copy to reliably capture the full table.
                _select_all_points_table(points_window, output_window, args.wait)

                parse_exc = None
                df = None
                for attempt in (1, 2):
                    previous_clip = pyperclip.paste()
                    copy_mode = _copy_current_curve_data(
                        points_window, output_window, args.copy_button, args.wait
                    )
                    table_text = _wait_for_clipboard_update(previous_clip, timeout_sec=1.5)
                    try:
                        df = parse_points_table(table_text).rename(
                            columns={"X": args.x_col, "Y": args.y_col}
                        )
                        break
                    except Exception as exc:
                        parse_exc = exc
                        # Retry once after refocusing/select-all.
                        _select_all_points_table(points_window, output_window, args.wait)
                        time.sleep(max(0.05, args.wait))
                        if attempt == 2:
                            raise parse_exc

                row_count = len(df)
                sheet_name = _unique_sheet_name(node_name, used_sheet_names)
                df.to_excel(writer, index=False, sheet_name=sheet_name)

                y_name = f"{safe_label(node_name)}_{args.y_col}"
                pair_df = df[[args.x_col, args.y_col]].rename(columns={args.y_col: y_name})
                if merged is None:
                    merged = pair_df.copy()
                else:
                    merged = merged.merge(pair_df, on=args.x_col, how="outer")

                success_count += 1
                status_rows.append(
                    {
                        "Node": node_name,
                        "Status": "OK",
                        "Rows": row_count,
                        "SelectMode": select_mode,
                        "CopyMode": copy_mode,
                        "Error": "",
                    }
                )
                logger(
                    f"Exported: {node_name} ({row_count} rows, "
                    f"select={select_mode}, copy={copy_mode})"
                )
            except Exception as exc:
                err_text = _error_text(exc)
                status_rows.append(
                    {
                        "Node": node_name,
                        "Status": "ERROR",
                        "Rows": row_count,
                        "SelectMode": select_mode,
                        "CopyMode": copy_mode,
                        "Error": err_text,
                    }
                )
                logger(f"Skipped: {node_name} ({err_text})")
                continue

        if merged is not None:
            merged.sort_values(args.x_col).to_excel(writer, index=False, sheet_name="Merged")
        pd.DataFrame(status_rows).to_excel(writer, index=False, sheet_name="_Status")

    if success_count == 0:
        first_error = ""
        for row in status_rows:
            if row["Status"] == "ERROR":
                first_error = row["Error"]
                break
        raise RuntimeError(
            "Spectrum export failed for all nodes. "
            f"Check _Status sheet in output file. First error: {first_error}"
        )
    logger(f"OK -> {out_path_final} ({success_count}/{len(node_names)} nodes)")


def nearest_node(nodes, x, y):
    best = None
    best_d2 = float("inf")
    for node in nodes:
        nx, ny = xy_of(node)
        d2 = (nx - x) ** 2 + (ny - y) ** 2
        if d2 < best_d2:
            best_d2 = d2
            best = node
    return best, best_d2 ** 0.5


def resolve_phase(g_o, phase_name, phase_index):
    if phase_name:
        for phase in g_o.Phases:
            try:
                p_name = str(phase.Identification.value)
            except Exception:
                p_name = str(phase)
            if (
                p_name == phase_name
                or p_name.startswith(phase_name + " ")
                or ("[" + phase_name + "]") in p_name
            ):
                return phase
        raise RuntimeError(f"Phase not found: {phase_name}")
    return g_o.Phases[phase_index]


def resolve_result_type(g_o, result_type_path):
    obj = g_o.ResultTypes
    for part in result_type_path.split("."):
        if not hasattr(obj, part):
            raise RuntimeError(f"Invalid result type path: {result_type_path}")
        obj = getattr(obj, part)
    return obj


def infer_related_result_type(result_type_path, target_prefix):
    parts = str(result_type_path).split(".")
    if not parts:
        return ""
    leaf = parts[-1].strip()
    if len(leaf) >= 2 and leaf[0].upper() in ("A", "V", "U"):
        parts[-1] = f"{target_prefix}{leaf[1:]}"
        return ".".join(parts)
    return ""


def _resolve_curve_time_result_type(g_o, acc_result_type_path, time_col):
    candidates = []
    prefix = ""
    if "." in str(acc_result_type_path):
        prefix = str(acc_result_type_path).rsplit(".", 1)[0].strip()

    raw_time = str(time_col or "").strip()
    if raw_time:
        if "." in raw_time:
            candidates.append(raw_time)
        elif prefix:
            candidates.append(f"{prefix}.{raw_time}")
        candidates.append(raw_time)
    if prefix:
        candidates.append(f"{prefix}.DynamicTime")
    for group_name in (
        "Soil",
        "Plate",
        "EmbeddedBeam",
        "FixedEndAnchor",
        "NodeToNodeAnchor",
    ):
        candidates.append(f"{group_name}.DynamicTime")

    seen = set()
    for candidate in candidates:
        key = str(candidate or "").strip()
        if not key or key in seen:
            continue
        seen.add(key)
        try:
            return resolve_result_type(g_o, key), key
        except Exception:
            continue
    return None, ""


def _safe_float_list(values):
    out = []
    for raw in values:
        try:
            out.append(float(raw))
        except Exception:
            out.append(float("nan"))
    return out


def _snap_dt(dt, target=0.01, rel_tol=0.02):
    try:
        dt_val = float(dt)
        target_val = float(target)
        tol_val = float(rel_tol)
    except Exception:
        return dt
    if dt_val <= 0.0 or target_val <= 0.0 or tol_val < 0.0:
        return dt
    if abs(dt_val - target_val) / target_val <= tol_val:
        return target_val
    return dt


def parse_points_argument(point_values):
    parsed = []
    for point_text in point_values:
        parts = [token.strip() for token in point_text.split(",")]
        if len(parts) != 2:
            raise RuntimeError(f'Invalid --point value "{point_text}", expected "x,y".')
        parsed.append((float(parts[0]), float(parts[1])))
    return parsed


def parse_points_file(path_text):
    if not path_text:
        return []
    path = Path(path_text)
    if not path.exists():
        raise RuntimeError(f"Points file not found: {path}")
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path)

    lower_map = {col.lower(): col for col in df.columns}
    if "x" in lower_map and "y" in lower_map:
        x_col = lower_map["x"]
        y_col = lower_map["y"]
    elif len(df.columns) >= 2:
        x_col = df.columns[0]
        y_col = df.columns[1]
    else:
        raise RuntimeError("Points file must contain at least two columns (x,y).")

    points = []
    for _, row in df[[x_col, y_col]].dropna().iterrows():
        points.append((float(row[x_col]), float(row[y_col])))
    return points


def run_timehistory_api(args, logger=print):
    s_o, g_o = _open_output_server(args.host, args.port, args.password)
    try:
        _ensure_output_result_types(g_o)
        _ensure_active_output_project(g_o)
        phase = resolve_phase(g_o, args.phase_name, args.phase_index)
        result_type = resolve_result_type(g_o, args.result_type)

        target_points = parse_points_argument(args.point) + parse_points_file(args.points_file)
        all_nodes = list(g_o.Nodes)

        node_records = []
        if args.all_nodes:
            for node in all_nodes:
                nx, ny = xy_of(node)
                node_records.append((nx, ny, nx, ny, 0.0))
        else:
            if not target_points:
                raise RuntimeError("Provide --point/--points-file or use --all-nodes.")
            for tx, ty in target_points:
                node, distance = nearest_node(all_nodes, tx, ty)
                nx, ny = xy_of(node)
                node_records.append((tx, ty, nx, ny, distance))

        steps = list(phase.Steps)
        if not steps:
            raise RuntimeError("Selected phase has no steps.")

        try:
            g_o.clearcurvepoints()
        except Exception:
            pass

        series_defs = []
        node_map_rows = []
        for idx, (tx, ty, nx, ny, distance) in enumerate(node_records, start=1):
            curve_point = g_o.addcurvepoint("node", nx, ny)
            col_name = f"N{idx}_{nx:.2f}_{ny:.2f}"
            series_defs.append((col_name, curve_point))
            node_map_rows.append(
                {
                    "Series": col_name,
                    "TargetX": tx,
                    "TargetY": ty,
                    "NodeX": nx,
                    "NodeY": ny,
                    "Distance": distance,
                }
            )
            logger(f"Prepared: {col_name}")

        rows = []
        value_cols = [name for name, _ in series_defs]
        valid_step_count = 0
        for idx, step in enumerate(steps, start=1):
            try:
                time_val = float(step.Reached.DynamicTime.value)
            except Exception:
                time_val = float(idx)

            row = {"Step": idx, args.time_col: time_val}
            valid_values = 0
            for col_name, curve_point in series_defs:
                try:
                    value = g_o.getsingleresult(step, result_type, curve_point)
                    row[col_name] = float(value)
                    valid_values += 1
                except Exception:
                    row[col_name] = float("nan")

            if valid_values > 0:
                valid_step_count += 1
            rows.append(row)

        history_df = pd.DataFrame(rows)
        history_df = history_df.dropna(subset=value_cols, how="all").reset_index(drop=True)
        if history_df.empty:
            raise RuntimeError(
                "No step-level values could be read for selected result type/phase. "
                "Check whether step results are available in Output."
            )
        if valid_step_count < len(steps):
            logger(
                f"Warning: values found in {valid_step_count}/{len(steps)} steps. "
                "Unavailable steps were omitted."
            )
        node_map_df = pd.DataFrame(node_map_rows)

        out_path = Path(args.out)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        if out_path.suffix.lower() == ".csv":
            out_path_final = _save_csv_with_fallback(history_df, out_path, logger=logger)
        else:
            writer, out_path_final = _open_excel_writer_with_fallback(out_path, logger=logger)
            with writer:
                history_df.to_excel(writer, index=False, sheet_name="TimeHistory")
                node_map_df.to_excel(writer, index=False, sheet_name="NodeMap")

        logger(f"OK -> {out_path_final}")
    finally:
        _safe_close_server(s_o)


def _read_curvepoint_metadata(curve_point, index):
    cp_id = str(getattr(curve_point, "_guid", f"cp_{index}"))
    try:
        identification = str(curve_point.Identification.value).strip()
    except Exception:
        identification = f"CurvePoint {index}"
    node_name = _match_node_name(identification) or identification
    try:
        x = float(curve_point.x.value)
        y = float(curve_point.y.value)
    except Exception:
        x = float("nan")
        y = float("nan")

    data_from = ""
    try:
        df = curve_point.DataFrom
        try:
            data_from = str(df.Name.value)
        except Exception:
            data_from = str(df)
    except Exception:
        data_from = ""

    label = f"{node_name} ({x:.2f},{y:.2f})"
    if data_from:
        label = f"{label} [{data_from}]"
    return {
        "id": cp_id,
        "index": index,
        "name": identification,
        "node_name": node_name,
        "x": x,
        "y": y,
        "data_from": data_from,
        "label": label,
        "obj": curve_point,
    }


def list_curve_points_api(host, port, password):
    s_o, g_o = _open_output_server(host, port, password)
    try:
        _ensure_output_result_types(g_o)
        _ensure_active_output_project(g_o)
        records = []
        for idx, cp in enumerate(list(g_o.CurvePoints), start=1):
            records.append(_read_curvepoint_metadata(cp, idx))
        return records
    finally:
        _safe_close_server(s_o)


def _make_unique_labels(labels):
    used = {}
    out = []
    for label in labels:
        key = safe_label(label) or "Series"
        count = used.get(key, 0) + 1
        used[key] = count
        if count == 1:
            out.append(key)
        else:
            out.append(f"{key}_{count}")
    return out


def _estimate_dt(time_values):
    arr = np.asarray(time_values, dtype=float)
    diffs = np.diff(arr)
    diffs = diffs[np.isfinite(diffs) & (diffs > 0)]
    if diffs.size == 0:
        raise RuntimeError("Could not estimate positive time step from DynamicTime.")
    return float(np.median(diffs))


def _compute_psa_spectrum(acc_series, dt, periods, damping):
    damping = float(damping)
    if damping < 0:
        raise RuntimeError("Damping ratio must be >= 0.")
    acc = np.asarray(acc_series, dtype=float)
    acc = np.where(np.isfinite(acc), acc, 0.0)
    periods = np.asarray(periods, dtype=float)
    psa = np.full(periods.shape, np.nan, dtype=float)
    if dt <= 0:
        return psa

    beta = 0.25
    gamma = 0.5

    for i, period in enumerate(periods):
        if not np.isfinite(period) or period <= 0:
            continue
        omega = 2.0 * math.pi / period
        k = omega * omega
        c = 2.0 * damping * omega

        a0 = 1.0 / (beta * dt * dt)
        a1 = gamma / (beta * dt)
        a2 = 1.0 / (beta * dt)
        a3 = (1.0 / (2.0 * beta)) - 1.0
        a4 = (gamma / beta) - 1.0
        a5 = dt * ((gamma / (2.0 * beta)) - 1.0)
        k_eff = k + a0 + a1 * c

        u = 0.0
        v = 0.0
        a = 0.0
        u_max = 0.0

        for ag in acc:
            p = -ag
            p_eff = p + (a0 * u + a2 * v + a3 * a) + c * (a1 * u + a4 * v + a5 * a)
            u_new = p_eff / k_eff
            a_new = a0 * (u_new - u) - a2 * v - a3 * a
            v_new = v + dt * ((1.0 - gamma) * a + gamma * a_new)

            u = u_new
            v = v_new
            a = a_new
            if abs(u) > u_max:
                u_max = abs(u)

        psa[i] = (omega * omega * u_max) / 9.81

    return psa


def run_curvepoints_api_export(args, logger=print):
    s_o, g_o = _open_output_server(args.host, args.port, args.password)
    try:
        _ensure_output_result_types(g_o)
        _ensure_active_output_project(g_o)
        phase = resolve_phase(g_o, args.phase_name, args.phase_index)
        acc_result_type_path = str(args.result_type).strip()
        if not acc_result_type_path:
            raise RuntimeError("Acceleration result type is required.")
        acc_result_type = resolve_result_type(g_o, acc_result_type_path)

        step_list = list(phase.Steps)
        if not step_list:
            raise RuntimeError("Selected phase has no steps.")

        all_records = []
        for idx, cp in enumerate(list(g_o.CurvePoints), start=1):
            all_records.append(_read_curvepoint_metadata(cp, idx))
        if not all_records:
            raise RuntimeError("No CurvePoints found. Save desired nodes to curve points first.")

        selected_ids = set(getattr(args, "curvepoint_id", []) or [])
        if selected_ids:
            selected = [r for r in all_records if r["id"] in selected_ids]
        else:
            selected = list(all_records)
        if not selected:
            raise RuntimeError("Selected CurvePoints not found in current Output session.")

        time_values = []
        for i, st in enumerate(step_list, start=1):
            try:
                time_values.append(float(st.Reached.DynamicTime.value))
            except Exception:
                time_values.append(float(i))
        step_count = len(step_list)

        base_labels = [rec["label"] for rec in selected]
        series_names = _make_unique_labels(base_labels)

        base_time_data = {
            "Step": list(range(1, step_count + 1)),
            args.time_col: time_values,
        }
        node_map_rows = []
        status_rows = []

        start_step = step_list[0]
        end_step = step_list[-1]

        for rec, col_name in zip(selected, series_names):
            node_map_rows.append(
                {
                    "CurvePointId": rec["id"],
                    "Series": col_name,
                    "CurvePointName": rec["name"],
                    "NodeName": rec["node_name"],
                    "X": rec["x"],
                    "Y": rec["y"],
                    "DataFrom": rec["data_from"],
                }
            )

        def read_history_by_result(result_type_obj, result_type_path, quantity):
            history_data = dict(base_time_data)
            value_cols = []
            rows = []
            for rec, col_name in zip(selected, series_names):
                try:
                    raw_vals = list(
                        g_o.getcurveresultspath([rec["obj"]], start_step, end_step, result_type_obj)
                    )
                    vals = []
                    for raw in raw_vals:
                        try:
                            vals.append(float(raw))
                        except Exception:
                            vals.append(float("nan"))
                    if len(vals) < step_count:
                        vals.extend([float("nan")] * (step_count - len(vals)))
                    elif len(vals) > step_count:
                        vals = vals[:step_count]

                    history_data[col_name] = vals
                    value_cols.append(col_name)
                    valid = int(np.isfinite(np.asarray(vals, dtype=float)).sum())
                    rows.append(
                        {
                            "Quantity": quantity,
                            "ResultType": result_type_path,
                            "CurvePointId": rec["id"],
                            "Series": col_name,
                            "Status": "OK",
                            "ValidCount": valid,
                            "Error": "",
                        }
                    )
                    logger(f"{quantity}: {col_name} ({valid}/{step_count})")
                except Exception as exc:
                    err = _error_text(exc)
                    rows.append(
                        {
                            "Quantity": quantity,
                            "ResultType": result_type_path,
                            "CurvePointId": rec["id"],
                            "Series": col_name,
                            "Status": "ERROR",
                            "ValidCount": 0,
                            "Error": err,
                        }
                    )
                    logger(f"{quantity} skipped: {col_name} ({err})")
                    continue

            if not value_cols:
                return None, [], rows

            history_df = pd.DataFrame(history_data)
            history_df = history_df.dropna(subset=value_cols, how="all").reset_index(drop=True)
            if history_df.empty:
                return None, [], rows

            return history_df, value_cols, rows

        def normalize_optional_result_type(raw_text, target_prefix):
            text = str(raw_text or "").strip()
            if text.lower() in ("none", "off", "-"):
                return None
            if text:
                return text
            inferred = infer_related_result_type(acc_result_type_path, target_prefix)
            return inferred or None

        velocity_path = normalize_optional_result_type(
            getattr(args, "velocity_result_type", ""), "V"
        )
        displacement_path = normalize_optional_result_type(
            getattr(args, "displacement_result_type", ""), "U"
        )

        logger(f"Acceleration result type: {acc_result_type_path}")
        if velocity_path:
            logger(f"Velocity result type: {velocity_path}")
        else:
            logger("Velocity result type: disabled")
        if displacement_path:
            logger(f"Displacement result type: {displacement_path}")
        else:
            logger("Displacement result type: disabled")

        history_df, value_cols, acc_status_rows = read_history_by_result(
            acc_result_type, acc_result_type_path, "Acceleration"
        )
        status_rows.extend(acc_status_rows)
        if not value_cols or history_df is None:
            raise RuntimeError(
                "No valid acceleration series could be read for selected phase/result type."
            )

        velocity_df = None
        if velocity_path:
            try:
                velocity_result_type = resolve_result_type(g_o, velocity_path)
                velocity_df, _, vel_status_rows = read_history_by_result(
                    velocity_result_type, velocity_path, "Velocity"
                )
                status_rows.extend(vel_status_rows)
                if velocity_df is None:
                    logger("Velocity history skipped (no valid rows).")
            except Exception as exc:
                err = _error_text(exc)
                status_rows.append(
                    {
                        "Quantity": "Velocity",
                        "ResultType": velocity_path,
                        "CurvePointId": "",
                        "Series": "",
                        "Status": "ERROR",
                        "ValidCount": 0,
                        "Error": err,
                    }
                )
                logger(f"Velocity history unavailable ({velocity_path}): {err}")

        displacement_df = None
        if displacement_path:
            try:
                displacement_result_type = resolve_result_type(g_o, displacement_path)
                displacement_df, _, disp_status_rows = read_history_by_result(
                    displacement_result_type, displacement_path, "Displacement"
                )
                status_rows.extend(disp_status_rows)
                if displacement_df is None:
                    logger("Displacement history skipped (no valid rows).")
            except Exception as exc:
                err = _error_text(exc)
                status_rows.append(
                    {
                        "Quantity": "Displacement",
                        "ResultType": displacement_path,
                        "CurvePointId": "",
                        "Series": "",
                        "Status": "ERROR",
                        "ValidCount": 0,
                        "Error": err,
                    }
                )
                logger(f"Displacement history unavailable ({displacement_path}): {err}")

        period_start = float(args.period_start)
        period_end = float(args.period_end)
        period_step = float(args.period_step)
        if period_start <= 0 or period_end <= 0 or period_step <= 0:
            raise RuntimeError("Period start/end/step must be > 0.")
        if period_end < period_start:
            raise RuntimeError("Period end must be >= period start.")

        periods = np.arange(period_start, period_end + period_step * 0.5, period_step, dtype=float)
        dt = _estimate_dt(history_df[args.time_col].to_numpy())
        logger(f"Spectrum settings: damping={args.damping}, dt~{dt:.6g}, periods={len(periods)}")

        spectrum_df = pd.DataFrame({"Period_s": periods})
        for col in value_cols:
            acc_vals = history_df[col].to_numpy(dtype=float)
            psa = _compute_psa_spectrum(acc_vals, dt, periods, float(args.damping))
            spectrum_df[col] = psa
            logger(f"Spectrum: {col} ({len(psa)} periods)")

        out_path = Path(args.out)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        node_map_df = pd.DataFrame(node_map_rows)
        status_df = pd.DataFrame(status_rows)
        split_files = bool(getattr(args, "split_output_files", False))

        if split_files:
            suffix = out_path.suffix if out_path.suffix.lower() == ".xlsx" else ".xlsx"
            if suffix != out_path.suffix:
                logger(
                    f"Split output requires Excel files; using .xlsx suffix instead of {out_path.suffix or '(none)'}."
                )
            base = out_path.with_suffix("")
            time_path = base.with_name(f"{base.name}_TimeHistory{suffix}")
            vel_path = base.with_name(f"{base.name}_VelocityHistory{suffix}")
            disp_path = base.with_name(f"{base.name}_DisplacementHistory{suffix}")
            spec_path = base.with_name(f"{base.name}_Spectrum{suffix}")
            map_path = base.with_name(f"{base.name}_NodeMap{suffix}")
            status_path = base.with_name(f"{base.name}_Status{suffix}")

            time_final = _write_single_sheet_xlsx_with_fallback(
                history_df, time_path, "TimeHistory", logger=logger
            )
            vel_final = None
            if velocity_df is not None:
                vel_final = _write_single_sheet_xlsx_with_fallback(
                    velocity_df, vel_path, "VelocityHistory", logger=logger
                )
            disp_final = None
            if displacement_df is not None:
                disp_final = _write_single_sheet_xlsx_with_fallback(
                    displacement_df, disp_path, "DisplacementHistory", logger=logger
                )
            spec_final = _write_single_sheet_xlsx_with_fallback(
                spectrum_df, spec_path, "Spectrum", logger=logger
            )
            map_final = _write_single_sheet_xlsx_with_fallback(
                node_map_df, map_path, "NodeMap", logger=logger
            )
            status_final = _write_single_sheet_xlsx_with_fallback(
                status_df, status_path, "_Status", logger=logger
            )
            logger(f"OK -> {time_final}")
            if vel_final is not None:
                logger(f"OK -> {vel_final}")
            if disp_final is not None:
                logger(f"OK -> {disp_final}")
            logger(f"OK -> {spec_final}")
            logger(f"OK -> {map_final}")
            logger(f"OK -> {status_final} ({len(value_cols)} series)")
        else:
            writer, out_path_final = _open_excel_writer_with_fallback(out_path, logger=logger)
            with writer:
                history_df.to_excel(writer, index=False, sheet_name="TimeHistory")
                if velocity_df is not None:
                    velocity_df.to_excel(writer, index=False, sheet_name="VelocityHistory")
                if displacement_df is not None:
                    displacement_df.to_excel(writer, index=False, sheet_name="DisplacementHistory")
                spectrum_df.to_excel(writer, index=False, sheet_name="Spectrum")
                node_map_df.to_excel(writer, index=False, sheet_name="NodeMap")
                status_df.to_excel(writer, index=False, sheet_name="_Status")

            logger(f"OK -> {out_path_final} ({len(value_cols)} series)")
    finally:
        _safe_close_server(s_o)


def _phase_display_name(phase):
    try:
        name = str(phase.Identification.value).strip()
    except Exception:
        name = str(phase).strip()
    return name or str(phase)


def _phase_short_name(phase_name):
    text = str(phase_name or "").strip()
    if not text:
        return text
    match = re.search(r"\[([^\]]+)\]\s*$", text)
    if match:
        return match.group(1).strip()
    return text


def _build_phase_alias_map(phases):
    phase_map = {}
    for phase in phases:
        display_name = _phase_display_name(phase)
        aliases = {display_name, _phase_short_name(display_name)}
        if " [" in display_name:
            aliases.add(display_name.split(" [", 1)[0].strip())
        for alias in aliases:
            alias = str(alias or "").strip()
            if alias and alias not in phase_map:
                phase_map[alias] = phase
    return phase_map


def _resolve_phase_by_name(phase_map, phase_name):
    key = str(phase_name or "").strip()
    if not key:
        raise RuntimeError("Phase name is empty.")
    if key in phase_map:
        return phase_map[key]
    short = _phase_short_name(key)
    if short in phase_map:
        return phase_map[short]
    raise RuntimeError(f"Phase not found: {phase_name}")


def _entity_name(entity):
    try:
        return str(entity.Name.value).strip()
    except Exception:
        return str(entity).strip()


def _entity_guid(entity):
    try:
        return str(getattr(entity, "_guid", ""))
    except Exception:
        return ""


def _entity_map_by_name(entities):
    out = {}
    for entity in entities:
        name = _entity_name(entity)
        if name and name not in out:
            out[name] = entity
    return out


def _safe_numeric_array(values, context):
    arr = np.asarray(list(values))
    if arr.size == 0:
        return np.asarray([], dtype=float)
    if arr.dtype.kind in ("U", "S", "O"):
        try:
            arr = np.asarray([float(v) for v in arr], dtype=float)
        except Exception:
            sample = "".join([str(v) for v in arr[:40]])
            raise RuntimeError(
                f"{context}: non-numeric results returned by PLAXIS ({sample[:120]})."
            )
    else:
        arr = arr.astype(float)
    return arr


def _get_results_numeric(g_o, obj, phase, result_type_obj, location, context):
    if obj is None:
        raw = g_o.getresults(phase, result_type_obj, location)
    else:
        raw = g_o.getresults(obj, phase, result_type_obj, location)
    return _safe_numeric_array(raw, context)


def _first_attr(obj, names):
    for name in names:
        try:
            value = getattr(obj, name)
        except Exception:
            continue
        if value is not None:
            return value, name
    return None, ""


def _resolve_structural_result_types(rt_group):
    candidates = {
        "M": (
            ["MEnvelopeMax2D", "M_EnvelopeMax2D", "MEnvelopeMax"],
            ["MEnvelopeMin2D", "M_EnvelopeMin2D", "MEnvelopeMin"],
        ),
        "N": (
            [
                "NEnvelopeMax2D",
                "N_EnvelopeMax2D",
                "NxEnvelopeMax2D",
                "Nx_EnvelopeMax2D",
                "NEnvelopeMax",
                "NxEnvelopeMax",
            ],
            [
                "NEnvelopeMin2D",
                "N_EnvelopeMin2D",
                "NxEnvelopeMin2D",
                "Nx_EnvelopeMin2D",
                "NEnvelopeMin",
                "NxEnvelopeMin",
            ],
        ),
        "Q": (
            ["QEnvelopeMax2D", "Q_EnvelopeMax2D", "QEnvelopeMax"],
            ["QEnvelopeMin2D", "Q_EnvelopeMin2D", "QEnvelopeMin"],
        ),
    }
    resolved = {}
    for key, (max_names, min_names) in candidates.items():
        max_rt, max_name = _first_attr(rt_group, max_names)
        min_rt, min_name = _first_attr(rt_group, min_names)
        resolved[key] = {
            "max": max_rt,
            "min": min_rt,
            "max_name": max_name,
            "min_name": min_name,
            "available": (max_rt is not None and min_rt is not None),
        }
    return resolved


def _find_geometry_phase(g_o):
    phases = list(g_o.Phases)
    for phase in phases:
        try:
            if list(phase.Steps):
                return phase
        except Exception:
            continue
    return phases[0] if phases else None


def _resolve_selected_curvepoints(all_records, selected_ids):
    selected_set = set(selected_ids or [])
    if selected_set:
        selected = [r for r in all_records if r["id"] in selected_set]
    else:
        selected = list(all_records)
    if not selected:
        raise RuntimeError("Selected CurvePoints were not found in current Output session.")
    return selected


def _prepare_periods(period_start, period_end, period_step):
    period_start = float(period_start)
    period_end = float(period_end)
    period_step = float(period_step)
    if period_start <= 0 or period_end <= 0 or period_step <= 0:
        raise RuntimeError("Period start/end/step must be > 0.")
    if period_end < period_start:
        raise RuntimeError("Period end must be >= period start.")
    return np.arange(period_start, period_end + period_step * 0.5, period_step, dtype=float)


def _ensure_plot_dir(base_out_path):
    out_path = Path(base_out_path)
    plot_dir = out_path.with_suffix("")
    plot_dir = plot_dir.parent / f"{plot_dir.name}_plots"
    plot_dir.mkdir(parents=True, exist_ok=True)
    return plot_dir


def _ensure_timehistory_dir(base_out_path):
    out_path = Path(base_out_path)
    hist_dir = out_path.with_suffix("")
    hist_dir = hist_dir.parent / f"{hist_dir.name}_time_history"
    hist_dir.mkdir(parents=True, exist_ok=True)
    return hist_dir


def _derive_output_with_suffix(base_out_path, suffix):
    out_path = Path(base_out_path)
    stemmed = out_path.with_suffix("")
    return stemmed.parent / f"{stemmed.name}_{suffix}.xlsx"


def _write_multisheet_workbook(out_path, sheets, logger=print):
    writer, out_final = _open_excel_writer_with_fallback(Path(out_path), logger=logger)
    with writer:
        for sheet_name, df in sheets:
            frame = df if isinstance(df, pd.DataFrame) else pd.DataFrame(df)
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
    return out_final


def _export_node_timehistory_subfolders(time_df, base_out_path, time_col, logger=print):
    if time_df.empty:
        return []

    root_dir = _ensure_timehistory_dir(base_out_path)
    written_paths = []
    key_rows = (
        time_df[["Direction", "Phase", "Series", "NodeName", "CurvePointId"]]
        .drop_duplicates()
        .itertuples(index=False)
    )
    for key in key_rows:
        direction, phase_name, series_name, node_name, curvepoint_id = key
        one = time_df[
            (time_df["Direction"] == direction)
            & (time_df["Phase"] == phase_name)
            & (time_df["Series"] == series_name)
            & (time_df["CurvePointId"] == curvepoint_id)
        ].copy()
        if one.empty:
            continue
        one = one.sort_values(time_col).reset_index(drop=True)
        phase_base = _safe_fs_name(_phase_base_name(phase_name))
        node_number = (
            _node_number(node_name)
            or _node_number(series_name)
            or _node_number(str(curvepoint_id))
        )
        if node_number:
            node_token = f"Node{node_number}"
        else:
            raw_node = str(node_name or series_name or curvepoint_id or "Node").strip()
            node_token = _safe_fs_name(raw_node.replace(" ", ""))

        node_dir = root_dir / node_token
        node_dir.mkdir(parents=True, exist_ok=True)
        txt_path = node_dir / f"{phase_base}.txt"
        export_df = one[[time_col, "Acceleration"]].copy()
        export_df.to_csv(txt_path, index=False, sep="\t")
        written_paths.append(str(txt_path))
        if logger is not None:
            logger(f"TimeHistory -> {txt_path}")
    return written_paths


def _build_node_spectrum_wide_specs(spectrum_long_df, spectrum_mean_df):
    if spectrum_long_df.empty:
        return []

    used_names = set(
        [
            sanitize_sheet_name("Phases"),
            sanitize_sheet_name("Selections"),
            sanitize_sheet_name("NodeTimeHistoryLong"),
            sanitize_sheet_name("NodeSpectrumLong"),
            sanitize_sheet_name("NodeSpectrumMean"),
            sanitize_sheet_name("_Status"),
        ]
    )
    specs = []
    directions = sorted(spectrum_long_df["Direction"].dropna().unique().tolist())

    for direction in directions:
        dir_long = spectrum_long_df[spectrum_long_df["Direction"] == direction].copy()
        dir_mean = spectrum_mean_df[spectrum_mean_df["Direction"] == direction].copy()

        for series in sorted(dir_long["Series"].dropna().unique().tolist()):
            one_long = dir_long[dir_long["Series"] == series].copy()
            if one_long.empty:
                continue
            wide = (
                one_long.pivot_table(
                    index="Period_s",
                    columns="Phase",
                    values="PSA_g",
                    aggfunc="mean",
                )
                .sort_index()
                .reset_index()
            )
            if not dir_mean.empty:
                one_mean = (
                    dir_mean[dir_mean["Series"] == series][["Period_s", "PSA_g"]]
                    .drop_duplicates(subset=["Period_s"])
                    .rename(columns={"PSA_g": "Mean"})
                )
                if not one_mean.empty:
                    wide = wide.merge(one_mean, on="Period_s", how="left")

            cols = wide.columns.tolist()
            phase_cols = sorted([c for c in cols if c not in ("Period_s", "Mean")])
            ordered = ["Period_s"] + phase_cols + (["Mean"] if "Mean" in cols else [])
            wide = wide[ordered]

            sheet_name = _unique_sheet_name(
                f"Spec_{direction}_{safe_label(series)}",
                used_names,
            )
            specs.append(
                {
                    "sheet_name": sheet_name,
                    "frame": wide,
                    "chart_title": f"{direction} | {series} | Spectrum",
                    "x_axis_title": "Period_s",
                    "y_axis_title": "PSA_g",
                    "chart_type": "scatter",
                    "shared_x_col": 1,
                }
            )

        phase_wide = (
            dir_long.groupby(["Period_s", "Phase"], as_index=False)
            .agg(PSA_g=("PSA_g", "mean"))
            .pivot(index="Period_s", columns="Phase", values="PSA_g")
            .sort_index()
            .reset_index()
        )
        if not phase_wide.empty:
            cols = phase_wide.columns.tolist()
            phase_cols = sorted([c for c in cols if c != "Period_s"])
            phase_wide = phase_wide[["Period_s"] + phase_cols]
            sheet_name = _unique_sheet_name(f"SpecPhase_{direction}", used_names)
            specs.append(
                {
                    "sheet_name": sheet_name,
                    "frame": phase_wide,
                    "chart_title": f"{direction} | Phase Overlay Mean (All Nodes)",
                    "x_axis_title": "Period_s",
                    "y_axis_title": "PSA_g",
                    "chart_type": "scatter",
                    "shared_x_col": 1,
                }
            )

        mean_wide = (
            dir_mean.pivot_table(
                index="Period_s",
                columns="Series",
                values="PSA_g",
                aggfunc="mean",
            )
            .sort_index()
            .reset_index()
        )
        if not mean_wide.empty:
            cols = mean_wide.columns.tolist()
            series_cols = sorted([c for c in cols if c != "Period_s"])
            mean_wide = mean_wide[["Period_s"] + series_cols]
            sheet_name = _unique_sheet_name(f"SpecMean_{direction}", used_names)
            specs.append(
                {
                    "sheet_name": sheet_name,
                    "frame": mean_wide,
                    "chart_title": f"{direction} | Node Mean Spectra Overlay",
                    "x_axis_title": "Period_s",
                    "y_axis_title": "PSA_g",
                    "chart_type": "scatter",
                    "shared_x_col": 1,
                }
            )

    return specs


def _build_stress_strain_wide_specs(stress_strain_df):
    if stress_strain_df.empty:
        return []

    used_names = set(
        [
            sanitize_sheet_name("Phases"),
            sanitize_sheet_name("Selections"),
            sanitize_sheet_name("StressStrainLong"),
            sanitize_sheet_name("_Status"),
        ]
    )
    specs = []
    directions = sorted(stress_strain_df["Direction"].dropna().unique().tolist())
    for direction in directions:
        dir_df = stress_strain_df[stress_strain_df["Direction"] == direction].copy()
        series_list = sorted(dir_df["Series"].dropna().unique().tolist())
        for series in series_list:
            sub = dir_df[dir_df["Series"] == series].copy()
            if sub.empty:
                continue

            frames = []
            series_pairs = []
            col_cursor = 1
            for phase_name in sorted(sub["Phase"].dropna().unique().tolist()):
                one = (
                    sub[sub["Phase"] == phase_name][["Gamma_xy", "Tau_xy"]]
                    .dropna(subset=["Gamma_xy", "Tau_xy"], how="any")
                    .reset_index(drop=True)
                )
                if one.empty:
                    continue
                phase_base = _phase_base_name(phase_name)
                x_col = f"{phase_base}_ShearStrain_pct"
                y_col = f"{phase_base}_ShearStress"
                one = one.assign(Gamma_xy_pct=pd.to_numeric(one["Gamma_xy"], errors="coerce") * 100.0)
                one = one.rename(columns={"Gamma_xy_pct": x_col, "Tau_xy": y_col})
                frames.append(one[[x_col, y_col]])
                series_pairs.append({"x_col": col_cursor, "y_col": col_cursor + 1, "title": phase_base})
                col_cursor += 2

            if not frames:
                continue

            wide = pd.concat(frames, axis=1)
            sheet_name = _unique_sheet_name(f"TauGam_{direction}_{safe_label(series)}", used_names)
            specs.append(
                {
                    "sheet_name": sheet_name,
                    "frame": wide,
                    "chart_title": f"{direction} | {series} | Shear Stress-Shear Strain",
                    "x_axis_title": "Shear Strain (%)",
                    "y_axis_title": "Shear Stress",
                    "chart_type": "scatter",
                    "series_pairs": series_pairs,
                    "chart_width": 18.0,
                    "chart_height": 8.5,
                    "x_num_fmt": "0.000",
                }
            )

    return specs


STRUCTURAL_COMPONENT_SPECS = (
    {
        "key": "M",
        "label": "Moment",
        "plus_col": "MPlus",
        "minus_col": "MMinus",
        "sheet_prefix": "MomentWide",
        "plot_prefix": "moment",
        "chart_title": "Moment-Profile Distance",
        "y_axis_title": "Moment (kNm)",
    },
    {
        "key": "N",
        "label": "Normal Force",
        "plus_col": "NPlus",
        "minus_col": "NMinus",
        "sheet_prefix": "NormalWide",
        "plot_prefix": "normal_force",
        "chart_title": "Normal Force-Profile Distance",
        "y_axis_title": "Normal Force",
    },
    {
        "key": "Q",
        "label": "Shear Force",
        "plus_col": "QPlus",
        "minus_col": "QMinus",
        "sheet_prefix": "ShearWide",
        "plot_prefix": "shear_force",
        "chart_title": "Shear Force-Profile Distance",
        "y_axis_title": "Shear Force",
    },
)


def _get_structural_component_spec(component_key):
    for spec in STRUCTURAL_COMPONENT_SPECS:
        if spec["key"] == component_key:
            return spec
    raise KeyError(f"Unknown structural component: {component_key}")


def _build_structural_component_wide_specs(avg_df, component_key):
    if avg_df.empty:
        return []

    spec = _get_structural_component_spec(component_key)
    plus_col = spec["plus_col"]
    minus_col = spec["minus_col"]
    used_names = set(
        [
            sanitize_sheet_name("Phases"),
            sanitize_sheet_name("Selections"),
            sanitize_sheet_name("MomentRawLong"),
            sanitize_sheet_name("MomentAvgByDir"),
            sanitize_sheet_name("_Status"),
        ]
    )
    specs = []
    directions = sorted(avg_df["Direction"].dropna().unique().tolist())
    for direction in directions:
        dir_df = avg_df[avg_df["Direction"] == direction].copy()
        groups = sorted(dir_df["ObjectGroup"].dropna().unique().tolist())
        for object_group in groups:
            sub = dir_df[dir_df["ObjectGroup"] == object_group].copy()
            if sub.empty:
                continue
            frames = []
            series_pairs = []
            col_cursor = 1
            object_names = sorted(sub["ObjectName"].dropna().unique().tolist())
            for object_name in object_names:
                one = (
                    sub[sub["ObjectName"] == object_name][["Depth", plus_col, minus_col]]
                    .dropna(subset=["Depth"])
                    .dropna(subset=[plus_col, minus_col], how="all")
                    .sort_values("Depth")
                    .reset_index(drop=True)
                )
                if one.empty:
                    continue
                c_depth = f"{object_name}_Depth"
                c_plus = f"{object_name}_{plus_col}"
                c_minus = f"{object_name}_{minus_col}"
                one = one.rename(columns={"Depth": c_depth, plus_col: c_plus, minus_col: c_minus})
                frames.append(one[[c_depth, c_plus, c_minus]])
                series_pairs.append(
                    {
                        "x_col": col_cursor,
                        "y_col": col_cursor + 1,
                        "title": c_plus,
                    }
                )
                series_pairs.append(
                    {
                        "x_col": col_cursor,
                        "y_col": col_cursor + 2,
                        "title": c_minus,
                    }
                )
                col_cursor += 3

            if not frames:
                continue

            wide = pd.concat(frames, axis=1)
            sheet_name = _unique_sheet_name(
                f"{spec['sheet_prefix']}_{direction}_{object_group}",
                used_names,
            )
            specs.append(
                {
                    "sheet_name": sheet_name,
                    "frame": wide,
                    "chart_title": f"{direction} | {object_group} | {spec['chart_title']}",
                    "x_axis_title": "Profile Distance (m)",
                    "y_axis_title": spec["y_axis_title"],
                    "chart_type": "scatter",
                    "series_pairs": series_pairs,
                    "chart_width": 18.5,
                    "chart_height": 9.5,
                }
            )

    return specs


def _add_excel_line_charts(workbook_path, chart_specs, logger=print):
    if not chart_specs:
        return

    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, Reference, ScatterChart, Series
    from openpyxl.chart.axis import ChartLines
    from openpyxl.utils import get_column_letter

    wb = load_workbook(workbook_path)
    changed = False
    for spec in chart_specs:
        sheet_name = spec.get("sheet_name")
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        max_row = int(ws.max_row or 0)
        max_col = int(ws.max_column or 0)
        if max_row < 2 or max_col < 2:
            continue

        chart_type = str(spec.get("chart_type") or "scatter").strip().lower()
        if chart_type == "line":
            chart = LineChart()
        else:
            chart = ScatterChart()
            chart.scatterStyle = "line"
        chart.title = str(spec.get("chart_title") or sheet_name)
        chart.style = 2
        chart.y_axis.title = str(spec.get("y_axis_title") or "Value")
        chart.x_axis.title = str(spec.get("x_axis_title") or "X")
        chart.height = float(spec.get("chart_height") or 9.0)
        chart.width = float(spec.get("chart_width") or 16.5)
        chart.legend.position = "r"
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.numFmt = str(spec.get("x_num_fmt") or "0.00")
        chart.y_axis.numFmt = str(spec.get("y_num_fmt") or "0.000")
        chart.x_axis.majorGridlines = ChartLines()
        chart.y_axis.majorGridlines = ChartLines()
        if bool(spec.get("reverse_x_axis", False)):
            try:
                chart.x_axis.scaling.orientation = "maxMin"
            except Exception:
                pass

        if chart_type == "line":
            data = Reference(ws, min_col=2, min_row=1, max_col=max_col, max_row=max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
        else:
            series_pairs = list(spec.get("series_pairs") or [])
            shared_x_col = int(spec.get("shared_x_col") or 0)
            if series_pairs:
                for pair in series_pairs:
                    x_col = int(pair["x_col"])
                    y_col = int(pair["y_col"])
                    xvalues = Reference(ws, min_col=x_col, min_row=2, max_row=max_row)
                    yvalues = Reference(ws, min_col=y_col, min_row=1, max_row=max_row)
                    series = Series(yvalues, xvalues, title_from_data=True)
                    chart.series.append(series)
            elif shared_x_col > 0:
                xvalues = Reference(ws, min_col=shared_x_col, min_row=2, max_row=max_row)
                for y_col in range(shared_x_col + 1, max_col + 1):
                    yvalues = Reference(ws, min_col=y_col, min_row=1, max_row=max_row)
                    series = Series(yvalues, xvalues, title_from_data=True)
                    chart.series.append(series)

        anchor_col = min(max_col + 2, 50)
        ws.add_chart(chart, f"{get_column_letter(anchor_col)}2")
        changed = True

    if changed:
        wb.save(workbook_path)
        if logger is not None:
            logger("Excel charts added to wide data sheets.")


def _mpl_pyplot():
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    return plt


def _phase_base_name(phase_name):
    text = str(phase_name or "").strip()
    if " [" in text:
        return text.split(" [", 1)[0].strip()
    return text


def _safe_fs_name(name):
    text = re.sub(r'[<>:"/\\|?*]+', "_", str(name or "").strip())
    text = text.strip(" .")
    return text or "item"


def _short_plot_label(text, max_len=40):
    out = str(text or "").strip()
    if len(out) <= max_len:
        return out
    return f"{out[: max_len - 3].rstrip()}..."


def _apply_compact_legend(fig, ax, prefer_right=False, right_rect=0.82):
    handles, labels = ax.get_legend_handles_labels()
    items = [(h, l) for h, l in zip(handles, labels) if l and l != "_nolegend_"]
    if len(items) <= 1:
        fig.tight_layout()
        return

    handles = [item[0] for item in items]
    labels = [item[1] for item in items]
    if not prefer_right and len(labels) <= 6:
        ncol = min(3, max(1, math.ceil(len(labels) / 2)))
        ax.legend(
            handles,
            labels,
            loc="upper center",
            bbox_to_anchor=(0.5, -0.16),
            fontsize=7,
            ncol=ncol,
            frameon=True,
            columnspacing=0.9,
            handlelength=1.8,
            borderaxespad=0.3,
        )
        fig.tight_layout(rect=(0, 0.1, 1, 1))
        return

    ax.legend(
        handles,
        labels,
        loc="center left",
        bbox_to_anchor=(1.02, 0.5),
        fontsize=6.8,
        ncol=1,
        frameon=True,
        handlelength=1.6,
        borderaxespad=0.3,
    )
    fig.tight_layout(rect=(0, 0, right_rect, 1))


def _anchor_axes_at_zero(ax, zero_x=False, zero_y=False):
    ax.margins(x=0.0, y=0.0)
    if zero_x:
        right = ax.get_xlim()[1]
        ax.set_xlim(left=0.0, right=max(0.0, right))
    if zero_y:
        top = ax.get_ylim()[1]
        ax.set_ylim(bottom=0.0, top=max(0.0, top))


def _plot_structural_component_group(avg_df, direction, object_group, component_key, out_png, dpi=150):
    subset = avg_df[
        (avg_df["Direction"] == direction) & (avg_df["ObjectGroup"] == object_group)
    ].copy()
    if subset.empty:
        return False

    spec = _get_structural_component_spec(component_key)
    plus_col = spec["plus_col"]
    minus_col = spec["minus_col"]
    subset = subset.dropna(subset=[plus_col, minus_col], how="all")
    if subset.empty:
        return False

    plt = _mpl_pyplot()
    fig, ax = plt.subplots(figsize=(10, 7))
    names = sorted(subset["ObjectName"].dropna().unique().tolist())
    cmap = plt.cm.get_cmap("tab20", max(len(names), 1))

    for idx, name in enumerate(names):
        one = (
            subset[subset["ObjectName"] == name]
            .dropna(subset=[plus_col, minus_col], how="all")
            .sort_values("Depth")
        )
        if one.empty:
            continue
        color = cmap(idx % cmap.N)
        label_name = _short_plot_label(name, max_len=28)
        if one[plus_col].notna().any():
            ax.plot(
                one[plus_col],
                one["Depth"],
                color=color,
                linewidth=1.7,
                label=f"{label_name} {component_key}+",
            )
        if one[minus_col].notna().any():
            ax.plot(
                one[minus_col],
                one["Depth"],
                color=color,
                linewidth=1.3,
                linestyle="--",
                label=f"{label_name} {component_key}-",
            )

    ax.axvline(0.0, color="black", linewidth=0.8, alpha=0.7)
    ax.invert_yaxis()
    ax.set_xlabel(spec["y_axis_title"])
    ax.set_ylabel("Profile Distance (m)")
    ax.set_title(f"{direction} Direction | {object_group} | {spec['label']} Mean +/- Envelope")
    ax.grid(True, alpha=0.25)
    _apply_compact_legend(fig, ax)
    fig.savefig(out_png, dpi=int(dpi))
    plt.close(fig)
    return True


def _plot_node_timehistory(direction, series_name, frame, out_png, time_col, dpi=150):
    if frame.empty:
        return False
    plt = _mpl_pyplot()
    fig, ax = plt.subplots(figsize=(10, 6))
    phases = sorted(frame["Phase"].unique().tolist())
    for phase_name in phases:
        one = frame[frame["Phase"] == phase_name].sort_values(time_col)
        ax.plot(
            one[time_col],
            one["Acceleration"],
            linewidth=1.0,
            alpha=0.9,
            label=_short_plot_label(_phase_base_name(phase_name), max_len=36),
        )
    ax.set_xlabel(time_col)
    ax.set_ylabel("Acceleration")
    ax.set_title(f"{direction} | {series_name} | Acceleration-Time")
    ax.grid(True, alpha=0.25)
    _anchor_axes_at_zero(ax, zero_x=True, zero_y=False)
    _apply_compact_legend(fig, ax)
    fig.savefig(out_png, dpi=int(dpi))
    plt.close(fig)
    return True


def _plot_node_spectrum_single(direction, series_name, long_df, mean_df, out_png, dpi=150):
    if long_df.empty and mean_df.empty:
        return False
    plt = _mpl_pyplot()
    fig, ax = plt.subplots(figsize=(10, 6))
    if not long_df.empty:
        phases = sorted(long_df["Phase"].unique().tolist())
        for phase_name in phases:
            one = long_df[long_df["Phase"] == phase_name].sort_values("Period_s")
            ax.plot(
                one["Period_s"],
                one["PSA_g"],
                linewidth=1.0,
                alpha=0.55,
                label=_short_plot_label(_phase_base_name(phase_name), max_len=36),
            )
    if not mean_df.empty:
        one_mean = mean_df.sort_values("Period_s")
        ax.plot(
            one_mean["Period_s"],
            one_mean["PSA_g"],
            color="black",
            linewidth=2.2,
            label="Mean",
        )
    ax.set_xlabel("Period_s")
    ax.set_ylabel("PSA_g")
    ax.set_title(f"{direction} | {series_name} | Spectrum")
    ax.grid(True, alpha=0.25)
    _anchor_axes_at_zero(ax, zero_x=True, zero_y=True)
    _apply_compact_legend(fig, ax)
    fig.savefig(out_png, dpi=int(dpi))
    plt.close(fig)
    return True


def _plot_node_spectrum_group_phase(direction, long_df, out_png, dpi=150):
    subset = long_df[long_df["Direction"] == direction].copy()
    if subset.empty:
        return False
    plt = _mpl_pyplot()
    fig, ax = plt.subplots(figsize=(11, 7))
    seen_phase = set()
    keys = subset[["Series", "Phase"]].drop_duplicates().itertuples(index=False)
    for key in keys:
        series_name, phase_name = key
        one = subset[(subset["Series"] == series_name) & (subset["Phase"] == phase_name)]
        one = one.sort_values("Period_s")
        label = _phase_base_name(phase_name) if phase_name not in seen_phase else "_nolegend_"
        seen_phase.add(phase_name)
        ax.plot(one["Period_s"], one["PSA_g"], linewidth=0.9, alpha=0.4, label=label)
    ax.set_xlabel("Period_s")
    ax.set_ylabel("PSA_g")
    ax.set_title(f"{direction} | All Node Spectra (Phase-Based Overlay)")
    ax.grid(True, alpha=0.25)
    _anchor_axes_at_zero(ax, zero_x=True, zero_y=True)
    _apply_compact_legend(fig, ax)
    fig.savefig(out_png, dpi=int(dpi))
    plt.close(fig)
    return True


def _plot_node_spectrum_group_mean(direction, mean_df, out_png, dpi=150):
    subset = mean_df[mean_df["Direction"] == direction].copy()
    if subset.empty:
        return False
    plt = _mpl_pyplot()
    fig, ax = plt.subplots(figsize=(11, 7))
    names = sorted(subset["Series"].unique().tolist())
    for name in names:
        one = subset[subset["Series"] == name].sort_values("Period_s")
        ax.plot(
            one["Period_s"],
            one["PSA_g"],
            linewidth=1.4,
            label=_short_plot_label(name, max_len=28),
        )
    ax.set_xlabel("Period_s")
    ax.set_ylabel("PSA_g")
    ax.set_title(f"{direction} | Node Mean Spectra Overlay")
    ax.grid(True, alpha=0.25)
    _anchor_axes_at_zero(ax, zero_x=True, zero_y=True)
    _apply_compact_legend(fig, ax)
    fig.savefig(out_png, dpi=int(dpi))
    plt.close(fig)
    return True


def _plot_stress_strain_single(direction, series_name, frame, out_png, dpi=150):
    if frame.empty:
        return False
    plt = _mpl_pyplot()
    fig, (ax, legend_ax) = plt.subplots(
        ncols=2,
        figsize=(14, 6.5),
        gridspec_kw={"width_ratios": [5.8, 1.6], "wspace": 0.03},
    )
    for phase_name in sorted(frame["Phase"].dropna().unique().tolist()):
        one = (
            frame[frame["Phase"] == phase_name][["Gamma_xy", "Tau_xy"]]
            .dropna(subset=["Gamma_xy", "Tau_xy"], how="any")
            .reset_index(drop=True)
        )
        if one.empty:
            continue
        gamma_percent = pd.to_numeric(one["Gamma_xy"], errors="coerce") * 100.0
        ax.plot(
            gamma_percent,
            one["Tau_xy"],
            linewidth=1.0,
            alpha=0.85,
            label=_short_plot_label(_phase_base_name(phase_name), max_len=28),
        )
    ax.axhline(0.0, color="black", linewidth=0.8, alpha=0.4)
    ax.axvline(0.0, color="black", linewidth=0.8, alpha=0.4)
    ax.set_xlabel("Shear Strain (%)")
    ax.set_ylabel("Shear Stress")
    ax.set_title(f"{direction} | {series_name} | Shear Stress-Shear Strain")
    ax.grid(True, alpha=0.25)
    legend_ax.axis("off")
    handles, labels = ax.get_legend_handles_labels()
    items = [(h, l) for h, l in zip(handles, labels) if l and l != "_nolegend_"]
    if items:
        legend_ax.legend(
            [item[0] for item in items],
            [item[1] for item in items],
            loc="center left",
            fontsize=6.8,
            frameon=True,
            handlelength=1.6,
            borderaxespad=0.2,
            labelspacing=0.45,
        )
    fig.subplots_adjust(left=0.075, right=0.985, bottom=0.12, top=0.90, wspace=0.02)
    fig.savefig(out_png, dpi=int(dpi), bbox_inches="tight", pad_inches=0.15)
    plt.close(fig)
    return True


def _profile_distance_from_xy(x_vals, y_vals):
    x_arr = np.asarray(x_vals, dtype=float)
    y_arr = np.asarray(y_vals, dtype=float)
    n = min(len(x_arr), len(y_arr))
    if n == 0:
        return np.zeros(0, dtype=float)

    x_arr = x_arr[:n]
    y_arr = y_arr[:n]
    out = np.zeros(n, dtype=float)
    finite = np.isfinite(x_arr) & np.isfinite(y_arr)
    if not finite.any():
        return out

    pts = np.column_stack((x_arr[finite], y_arr[finite]))
    if len(pts) <= 1:
        out[finite] = 0.0
        return out

    centered = pts - pts.mean(axis=0)
    try:
        _, _, vh = np.linalg.svd(centered, full_matrices=False)
        axis = np.asarray(vh[0], dtype=float)
    except Exception:
        axis = np.array([1.0, 0.0], dtype=float)

    if (not np.all(np.isfinite(axis))) or np.allclose(axis, 0.0):
        spread_x = float(np.nanmax(pts[:, 0]) - np.nanmin(pts[:, 0]))
        spread_y = float(np.nanmax(pts[:, 1]) - np.nanmin(pts[:, 1]))
        axis = np.array([1.0, 0.0], dtype=float) if spread_x >= spread_y else np.array([0.0, 1.0], dtype=float)

    horizontal = abs(axis[0]) > abs(axis[1])
    if horizontal and axis[0] < 0.0:
        axis = -axis
    if (not horizontal) and axis[1] < 0.0:
        axis = -axis

    proj = pts @ axis
    if horizontal:
        dist = proj - float(np.nanmin(proj))
    else:
        dist = float(np.nanmax(proj)) - proj
    out[finite] = dist
    return out


def _collect_model_node_cloud(g_o, max_points=25000):
    points = []
    try:
        model_nodes = list(g_o.Nodes)
    except Exception:
        return points

    for node in model_nodes:
        try:
            x, y = xy_of(node)
        except Exception:
            continue
        if np.isfinite(x) and np.isfinite(y):
            points.append((float(x), float(y)))

    if not points:
        return points

    max_points = int(max_points or 0)
    if max_points > 0 and len(points) > max_points:
        step = max(1, int(math.ceil(len(points) / float(max_points))))
        points = points[::step]
    return points


def _build_node_export_context(g_o, selected_ids, primary_result_type_path):
    all_records = []
    for idx, cp in enumerate(list(g_o.CurvePoints), start=1):
        all_records.append(_read_curvepoint_metadata(cp, idx))
    if not all_records:
        raise RuntimeError("No CurvePoints found. Save desired nodes to curve points first.")

    selected = _resolve_selected_curvepoints(all_records, selected_ids)
    model_node_cloud = _collect_model_node_cloud(g_o)
    result_prefix = str(primary_result_type_path).split(".", 1)[0].strip().lower()
    known_data_from = 0
    mismatched_data_from = 0
    for rec in selected:
        data_from_text = str(rec.get("data_from") or "").strip()
        if not data_from_text:
            continue
        known_data_from += 1
        if not data_from_text.lower().startswith(result_prefix):
            mismatched_data_from += 1

    base_labels = [rec["label"] for rec in selected]
    series_names = _make_unique_labels(base_labels)
    series_by_id = {rec["id"]: name for rec, name in zip(selected, series_names)}

    selection_rows = []
    node_map_rows = []
    for rec in selected:
        selection_rows.append(
            {
                "SelectionType": "CurvePoint",
                "ObjectType": "Node",
                "ObjectName": rec["name"],
                "CurvePointId": rec["id"],
                "Series": series_by_id.get(rec["id"], rec["name"]),
                "NodeName": rec["node_name"],
                "X": rec["x"],
                "Y": rec["y"],
            }
        )
        node_map_rows.append(
            {
                "CurvePointId": rec["id"],
                "Series": series_by_id.get(rec["id"], rec["name"]),
                "CurvePointName": rec["name"],
                "NodeName": rec["node_name"],
                "X": rec["x"],
                "Y": rec["y"],
                "DataFrom": rec["data_from"],
            }
        )

    return {
        "all_records": all_records,
        "selected": selected,
        "model_node_cloud": model_node_cloud,
        "result_prefix": result_prefix,
        "known_data_from": known_data_from,
        "mismatched_data_from": mismatched_data_from,
        "series_by_id": series_by_id,
        "selection_rows": selection_rows,
        "node_map_rows": node_map_rows,
        "phase_map": _build_phase_alias_map(list(g_o.Phases)),
    }


def _resolve_phase_time_series(step_list):
    phase_t_values = []
    for st in step_list:
        try:
            phase_t_values.append(float(st.Reached.DynamicTime.value))
        except Exception:
            phase_t_values.append(float("nan"))

    phase_dt = None
    valid_times = [v for v in phase_t_values if np.isfinite(v)]
    if len(valid_times) >= 2:
        try:
            phase_dt = _estimate_dt(np.asarray(valid_times, dtype=float))
        except Exception:
            phase_dt = None
    if (phase_dt is None or phase_dt <= 0.0) and len(step_list) >= 2:
        try:
            t0 = float(step_list[0].Reached.DynamicTime.value)
            t1 = float(step_list[-1].Reached.DynamicTime.value)
            if np.isfinite(t0) and np.isfinite(t1) and t1 > t0:
                phase_dt = float((t1 - t0) / (len(step_list) - 1))
        except Exception:
            pass

    if phase_dt is not None and phase_dt > 0.0:
        base_start = 0.0
        for value in phase_t_values:
            if np.isfinite(value):
                base_start = float(value)
                break
        arr = np.asarray(phase_t_values, dtype=float)
        if (not np.all(np.isfinite(arr))) or np.any(np.diff(arr) <= 0.0):
            phase_t_values = [float(base_start + phase_dt * idx) for idx in range(len(step_list))]
    else:
        phase_t_values = [float(i) for i in range(1, len(step_list) + 1)]

    return phase_t_values, phase_dt


def _read_curve_time_values(
    g_o,
    curve_obj,
    start_step,
    end_step,
    curve_time_result_type,
    count,
    step_count,
    phase_t_values,
    phase_dt,
):
    time_vals = None
    if curve_time_result_type is not None:
        try:
            raw_t = list(
                g_o.getcurveresultspath(
                    [curve_obj],
                    start_step,
                    end_step,
                    curve_time_result_type,
                )
            )
            t_vals = _safe_float_list(raw_t)
            if len(t_vals) == count:
                time_vals = t_vals
        except Exception:
            time_vals = None

    if time_vals is None:
        if step_count >= count and step_count > 0:
            time_vals = phase_t_values[:count]
        elif phase_dt is not None and phase_dt > 0.0:
            start_t = phase_t_values[0] if phase_t_values else 0.0
            time_vals = [start_t + phase_dt * i for i in range(count)]
        else:
            time_vals = [float(i) for i in range(1, count + 1)]
    return time_vals


def _plot_node_selection_map(model_points, node_map_df, out_png, dpi=150):
    if node_map_df.empty:
        return False

    sel = node_map_df.copy()
    sel["X"] = pd.to_numeric(sel["X"], errors="coerce")
    sel["Y"] = pd.to_numeric(sel["Y"], errors="coerce")
    sel = sel[np.isfinite(sel["X"]) & np.isfinite(sel["Y"])].copy()
    if sel.empty:
        return False

    plt = _mpl_pyplot()
    fig, ax = plt.subplots(figsize=(10, 6))

    if model_points:
        bg_x, bg_y = zip(*model_points)
        ax.scatter(
            bg_x,
            bg_y,
            s=5,
            c="#B8BEC7",
            alpha=0.35,
            edgecolors="none",
            rasterized=True,
            label="PLAXIS model nodes",
        )

    ax.scatter(
        sel["X"],
        sel["Y"],
        s=52,
        c="#D62728",
        edgecolors="black",
        linewidths=0.5,
        zorder=3,
        label="Selected",
    )

    for row in sel.itertuples(index=False):
        label = str(getattr(row, "Series", "") or getattr(row, "NodeName", "") or "").strip()
        if not label:
            label = str(getattr(row, "CurvePointId", "")).strip()
        label = _short_plot_label(label, max_len=28)
        ax.annotate(
            label,
            (row.X, row.Y),
            xytext=(6, 6),
            textcoords="offset points",
            fontsize=8,
            zorder=4,
        )

    ax.set_xlabel("X")
    ax.set_ylabel("Y")
    ax.set_title("Selected CurvePoints | Model XY View")
    ax.grid(True, alpha=0.25)
    ax.set_aspect("equal", adjustable="box")
    _apply_compact_legend(fig, ax)
    fig.savefig(out_png, dpi=int(dpi))
    plt.close(fig)
    return True


def _apply_plate_group_merge(raw_df, merge_group1=False, merge_group2=False):
    if raw_df.empty:
        return raw_df

    out = raw_df.copy()
    merge_specs = []
    if merge_group1:
        merge_specs.append(("PlateGroup1", "PlateGroup1_Merged"))
    if merge_group2:
        merge_specs.append(("PlateGroup2", "PlateGroup2_Merged"))

    for group_name, merged_name in merge_specs:
        mask = out["ObjectGroup"] == group_name
        if not mask.any():
            continue

        out.loc[mask, "ObjectName"] = merged_name
        key_cols = ["Direction", "Phase", "ObjectGroup", "ObjectName"]
        grouped = out.loc[mask].groupby(key_cols, sort=False).groups
        for row_index in grouped.values():
            coords = out.loc[row_index, ["X", "Y"]]
            out.loc[row_index, "Depth"] = _profile_distance_from_xy(
                coords["X"].to_numpy(),
                coords["Y"].to_numpy(),
            )

    return out


def list_phases_api(host, port, password):
    s_o, g_o = _open_output_server(host, port, password)
    try:
        _ensure_active_output_project(g_o)
        records = []
        for idx, phase in enumerate(list(g_o.Phases), start=1):
            name = _phase_display_name(phase)
            records.append(
                {
                    "index": idx,
                    "id": _entity_guid(phase),
                    "name": name,
                    "short_name": _phase_short_name(name),
                }
            )
        return records
    finally:
        _safe_close_server(s_o)


def list_structural_objects_api(host, port, password):
    s_o, g_o = _open_output_server(host, port, password)
    try:
        _ensure_output_result_types(g_o)
        _ensure_active_output_project(g_o)
        phase = _find_geometry_phase(g_o)
        out = {"embedded_beams": [], "plates": []}

        def build_record_list(objects, rt_group, key):
            records = []
            for idx, obj in enumerate(objects, start=1):
                name = _entity_name(obj)
                rec = {"index": idx, "id": _entity_guid(obj), "name": name, "label": name}
                if phase is not None:
                    try:
                        x = _get_results_numeric(
                            g_o, obj, phase, rt_group.X, "node", f"{key}:{name}:X"
                        )
                        y = _get_results_numeric(
                            g_o, obj, phase, rt_group.Y, "node", f"{key}:{name}:Y"
                        )
                        if x.size and y.size:
                            rec["x_min"] = float(np.nanmin(x))
                            rec["x_max"] = float(np.nanmax(x))
                            rec["y_min"] = float(np.nanmin(y))
                            rec["y_max"] = float(np.nanmax(y))
                            rec["label"] = (
                                f"{name} | x={rec['x_min']:.2f}..{rec['x_max']:.2f} "
                                f"| y={rec['y_min']:.2f}..{rec['y_max']:.2f}"
                            )
                    except Exception:
                        pass
                records.append(rec)
            return records

        out["embedded_beams"] = build_record_list(
            list(g_o.EmbeddedBeams), g_o.ResultTypes.EmbeddedBeam, "EmbeddedBeam"
        )
        out["plates"] = build_record_list(list(g_o.Plates), g_o.ResultTypes.Plate, "Plate")
        return out
    finally:
        _safe_close_server(s_o)


def run_structural_moment_export(args, logger=print):
    x_phase_names = list(getattr(args, "x_phase_names", []) or [])
    y_phase_names = list(getattr(args, "y_phase_names", []) or [])
    pile_names = list(getattr(args, "embedded_beam_names", []) or [])
    plate_group1 = list(getattr(args, "plate_group1_names", []) or [])
    plate_group2 = list(getattr(args, "plate_group2_names", []) or [])
    merge_plate_group1 = bool(
        getattr(args, "plate_group1_merge_single_profile", False)
    )
    merge_plate_group2 = bool(
        getattr(args, "plate_group2_merge_single_profile", False)
    )
    plot_dpi = int(float(getattr(args, "plot_dpi", 150) or 150))
    out_text = str(getattr(args, "out", "")).strip()
    out_path = Path(out_text).expanduser()
    if not str(getattr(args, "password", "")).strip():
        raise RuntimeError("Password is required.")
    if not x_phase_names and not y_phase_names:
        raise RuntimeError("Select at least one X or Y phase.")
    if not pile_names and not plate_group1 and not plate_group2:
        raise RuntimeError("Select at least one pile or plate object.")
    if not out_text:
        raise RuntimeError("Output path is required.")

    s_o, g_o = _open_output_server(args.host, int(args.port), args.password)
    try:
        _ensure_output_result_types(g_o)
        _ensure_active_output_project(g_o)
        phase_list = list(g_o.Phases)
        phase_map = _build_phase_alias_map(phase_list)

        pile_map = _entity_map_by_name(list(g_o.EmbeddedBeams))
        plate_map = _entity_map_by_name(list(g_o.Plates))

        selected_groups = [
            {
                "object_group": "Pile",
                "object_type": "EmbeddedBeam",
                "names": pile_names,
                "obj_map": pile_map,
                "rt_group": g_o.ResultTypes.EmbeddedBeam,
            },
            {
                "object_group": "PlateGroup1",
                "object_type": "Plate",
                "names": plate_group1,
                "obj_map": plate_map,
                "rt_group": g_o.ResultTypes.Plate,
            },
            {
                "object_group": "PlateGroup2",
                "object_type": "Plate",
                "names": plate_group2,
                "obj_map": plate_map,
                "rt_group": g_o.ResultTypes.Plate,
            },
        ]
        for group in selected_groups:
            group["resolved_components"] = _resolve_structural_result_types(group["rt_group"])

        phases_rows = []
        selection_rows = []
        status_rows = []
        raw_rows = []

        total_phase_count = len(x_phase_names) + len(y_phase_names)
        logger(
            f"Structural run started: phases={total_phase_count}, "
            f"piles={len(pile_names)}, plate_g1={len(plate_group1)}, plate_g2={len(plate_group2)}, "
            f"merge_g1={int(merge_plate_group1)}, merge_g2={int(merge_plate_group2)}"
        )
        for group in selected_groups:
            names = list(group["names"])
            if not names:
                continue
            missing = [
                _get_structural_component_spec(key)["label"]
                for key, info in group["resolved_components"].items()
                if not info["available"]
            ]
            if missing:
                message = (
                    f"{group['object_group']} result types not found for: {', '.join(missing)}. "
                    "Those outputs will be skipped."
                )
                logger(f"Warning: {message}")
                status_rows.append(
                    {
                        "Category": "ResultTypeResolve",
                        "Direction": "",
                        "Phase": "",
                        "ObjectGroup": group["object_group"],
                        "ObjectName": "",
                        "Status": "WARN",
                        "Message": message,
                    }
                )
        phase_counter = 0
        for direction, phase_names in (("X", x_phase_names), ("Y", y_phase_names)):
            for phase_name in phase_names:
                phase_counter += 1
                try:
                    phase = _resolve_phase_by_name(phase_map, phase_name)
                except Exception as exc:
                    status_rows.append(
                        {
                            "Category": "PhaseResolve",
                            "Direction": direction,
                            "Phase": phase_name,
                            "ObjectGroup": "",
                            "ObjectName": "",
                            "Status": "ERROR",
                            "Message": _error_text(exc),
                        }
                    )
                    continue

                resolved_phase_name = _phase_display_name(phase)
                logger(f"[{phase_counter}/{total_phase_count}] {direction} phase: {resolved_phase_name}")
                phases_rows.append({"Direction": direction, "Phase": resolved_phase_name})
                for group in selected_groups:
                    object_group = group["object_group"]
                    object_type = group["object_type"]
                    obj_map = group["obj_map"]
                    resolved_components = group["resolved_components"]
                    for object_name in group["names"]:
                        selection_rows.append(
                            {
                                "SelectionType": object_group,
                                "ObjectType": object_type,
                                "ObjectName": object_name,
                            }
                        )
                        obj = obj_map.get(object_name)
                        if obj is None:
                            status_rows.append(
                                {
                                    "Category": "ObjectResolve",
                                    "Direction": direction,
                                    "Phase": resolved_phase_name,
                                    "ObjectGroup": object_group,
                                    "ObjectName": object_name,
                                    "Status": "ERROR",
                                    "Message": "Object not found in current Output session.",
                                }
                            )
                            continue

                        context = f"{direction}:{resolved_phase_name}:{object_group}:{object_name}"
                        try:
                            x = _get_results_numeric(
                                g_o, obj, phase, group["rt_group"].X, "node", f"{context}:X"
                            )
                            y = _get_results_numeric(
                                g_o, obj, phase, group["rt_group"].Y, "node", f"{context}:Y"
                            )
                            component_values = {}
                            component_names = []
                            lengths = [len(x), len(y)]
                            for component_key in ("M", "N", "Q"):
                                info = resolved_components.get(component_key, {})
                                if not info.get("available"):
                                    component_values[component_key] = (None, None)
                                    continue
                                plus_vals = _get_results_numeric(
                                    g_o,
                                    obj,
                                    phase,
                                    info["max"],
                                    "node",
                                    f"{context}:{component_key}+",
                                )
                                minus_vals = _get_results_numeric(
                                    g_o,
                                    obj,
                                    phase,
                                    info["min"],
                                    "node",
                                    f"{context}:{component_key}-",
                                )
                                component_values[component_key] = (plus_vals, minus_vals)
                                component_names.append(component_key)
                                lengths.extend([len(plus_vals), len(minus_vals)])
                            if not component_names:
                                raise RuntimeError("No force/moment envelope result types available.")
                            n = min(lengths) if lengths else 0
                            if n == 0:
                                raise RuntimeError("No numeric node results.")
                            x = x[:n]
                            y = y[:n]
                            trimmed_components = {}
                            for component_key, pair in component_values.items():
                                if pair[0] is None or pair[1] is None:
                                    trimmed_components[component_key] = (np.nan, np.nan)
                                    continue
                                trimmed_components[component_key] = (pair[0][:n], pair[1][:n])
                            depth = _profile_distance_from_xy(x, y)

                            for i in range(n):
                                m_plus, m_minus = trimmed_components["M"]
                                n_plus, n_minus = trimmed_components["N"]
                                q_plus, q_minus = trimmed_components["Q"]
                                raw_rows.append(
                                    {
                                        "Direction": direction,
                                        "Phase": resolved_phase_name,
                                        "ObjectGroup": object_group,
                                        "ObjectType": object_type,
                                        "ObjectName": object_name,
                                        "X": float(x[i]),
                                        "Y": float(y[i]),
                                        "Depth": float(depth[i]),
                                        "MPlus": float(m_plus[i]) if np.ndim(m_plus) else np.nan,
                                        "MMinus": float(m_minus[i]) if np.ndim(m_minus) else np.nan,
                                        "NPlus": float(n_plus[i]) if np.ndim(n_plus) else np.nan,
                                        "NMinus": float(n_minus[i]) if np.ndim(n_minus) else np.nan,
                                        "QPlus": float(q_plus[i]) if np.ndim(q_plus) else np.nan,
                                        "QMinus": float(q_minus[i]) if np.ndim(q_minus) else np.nan,
                                    }
                                )
                            status_rows.append(
                                {
                                    "Category": "StructuralRead",
                                    "Direction": direction,
                                    "Phase": resolved_phase_name,
                                    "ObjectGroup": object_group,
                                    "ObjectName": object_name,
                                    "Status": "OK",
                                    "Message": f"{n} points | components={','.join(component_names) or 'none'}",
                                }
                            )
                            logger(
                                f"{direction} {resolved_phase_name} | "
                                f"{object_group}:{object_name} -> {n} points | "
                                f"components={','.join(component_names) or 'none'}"
                            )
                        except Exception as exc:
                            status_rows.append(
                                {
                                    "Category": "StructuralRead",
                                    "Direction": direction,
                                    "Phase": resolved_phase_name,
                                    "ObjectGroup": object_group,
                                    "ObjectName": object_name,
                                    "Status": "ERROR",
                                    "Message": _error_text(exc),
                                }
                            )

        raw_df = pd.DataFrame(raw_rows)
        if raw_df.empty:
            raise RuntimeError(
                "No structural force/moment data could be read for selected phases/objects."
            )

        raw_df = _apply_plate_group_merge(
            raw_df,
            merge_group1=merge_plate_group1,
            merge_group2=merge_plate_group2,
        )
        raw_df = raw_df.sort_values(
            ["Direction", "ObjectGroup", "ObjectName", "Phase", "Depth"]
        ).reset_index(drop=True)
        if merge_plate_group1:
            logger("PlateGroup1 merge mode enabled: selected plates treated as one profile.")
        if merge_plate_group2:
            logger("PlateGroup2 merge mode enabled: selected plates treated as one profile.")

        work = raw_df.copy()
        work["DepthRound"] = work["Depth"].round(6)
        avg_df = (
            work.groupby(
                ["Direction", "ObjectGroup", "ObjectType", "ObjectName", "DepthRound"],
                as_index=False,
            )
            .agg(
                Depth=("Depth", "mean"),
                MPlus=("MPlus", "mean"),
                MMinus=("MMinus", "mean"),
                NPlus=("NPlus", "mean"),
                NMinus=("NMinus", "mean"),
                QPlus=("QPlus", "mean"),
                QMinus=("QMinus", "mean"),
                SampleCount=("Phase", "count"),
            )
            .sort_values(["Direction", "ObjectGroup", "ObjectName", "Depth"])
            .reset_index(drop=True)
        )

        out_path.parent.mkdir(parents=True, exist_ok=True)
        plot_dir = _ensure_plot_dir(out_path)
        plot_files = []
        for component_key in ("M", "N", "Q"):
            spec = _get_structural_component_spec(component_key)
            for direction in ("X", "Y"):
                for object_group in ("Pile", "PlateGroup1", "PlateGroup2"):
                    if avg_df[
                        (avg_df["Direction"] == direction) & (avg_df["ObjectGroup"] == object_group)
                    ].empty:
                        continue
                    png_path = plot_dir / f"{spec['plot_prefix']}_{direction}_{object_group}.png"
                    ok = _plot_structural_component_group(
                        avg_df,
                        direction,
                        object_group,
                        component_key,
                        png_path,
                        dpi=plot_dpi,
                    )
                    if ok:
                        plot_files.append(str(png_path))
                        logger(f"Chart -> {png_path}")

        for path_text in plot_files:
            status_rows.append(
                {
                    "Category": "Chart",
                    "Direction": "",
                    "Phase": "",
                    "ObjectGroup": "",
                    "ObjectName": "",
                    "Status": "OK",
                    "Message": path_text,
                }
            )

        phases_df = pd.DataFrame(phases_rows).drop_duplicates()
        selections_df = pd.DataFrame(selection_rows).drop_duplicates()
        status_df = pd.DataFrame(status_rows)
        structural_wide_specs = []
        for component_key in ("M", "N", "Q"):
            structural_wide_specs.extend(_build_structural_component_wide_specs(avg_df, component_key))

        sheets = [
            ("Phases", phases_df),
            ("Selections", selections_df),
            ("MomentRawLong", raw_df),
            ("MomentAvgByDir", avg_df),
            ("_Status", status_df),
        ]
        for spec in structural_wide_specs:
            sheets.append((spec["sheet_name"], spec["frame"]))
        out_final = _write_multisheet_workbook(out_path, sheets, logger=logger)
        _add_excel_line_charts(out_final, structural_wide_specs, logger=logger)
        logger(f"OK -> {out_final}")
        logger(f"Charts -> {plot_dir}")
    finally:
        _safe_close_server(s_o)


def run_node_multiphase_spectrum_export(args, logger=print):
    x_phase_names = list(getattr(args, "x_phase_names", []) or [])
    y_phase_names = list(getattr(args, "y_phase_names", []) or [])
    selected_ids = list(getattr(args, "curvepoint_id", []) or [])
    out_text = str(getattr(args, "out", "")).strip()
    out_path = Path(out_text).expanduser()
    plot_dpi = int(float(getattr(args, "plot_dpi", 150) or 150))
    save_node_timehistory_subfolders = bool(
        getattr(args, "save_node_timehistory_subfolders", False)
    )
    if not str(getattr(args, "password", "")).strip():
        raise RuntimeError("Password is required.")
    if not x_phase_names and not y_phase_names:
        raise RuntimeError("Select at least one X or Y phase.")
    if not out_text:
        raise RuntimeError("Output path is required.")

    s_o, g_o = _open_output_server(args.host, int(args.port), args.password)
    try:
        _ensure_output_result_types(g_o)
        _ensure_active_output_project(g_o)
        acc_result_type_path = str(args.result_type).strip()
        if not acc_result_type_path:
            raise RuntimeError("Acceleration result type is required.")
        acc_result_type = resolve_result_type(g_o, acc_result_type_path)
        periods = _prepare_periods(args.period_start, args.period_end, args.period_step)

        ctx = _build_node_export_context(g_o, selected_ids, acc_result_type_path)
        selected = ctx["selected"]
        series_by_id = ctx["series_by_id"]
        phase_map = ctx["phase_map"]
        phases_rows = []
        selection_rows = list(ctx["selection_rows"])
        node_map_rows = list(ctx["node_map_rows"])
        status_rows = []
        time_rows = []
        spectrum_rows = []

        time_col = str(getattr(args, "time_col", "DynamicTime")).strip() or "DynamicTime"
        damping = float(args.damping)
        dt_snap_target = float(getattr(args, "dt_snap_target", 0.01) or 0.0)
        dt_snap_rel_tol = float(getattr(args, "dt_snap_rel_tol", 0.02) or 0.0)
        curve_time_result_type, curve_time_result_type_path = _resolve_curve_time_result_type(
            g_o,
            acc_result_type_path,
            time_col,
        )

        total_phase_count = len(x_phase_names) + len(y_phase_names)
        logger(
            f"Node spectrum run started: phases={total_phase_count}, "
            f"curvepoints={len(selected)}, result_type={acc_result_type_path}, "
            f"plot_dpi={plot_dpi}"
        )
        if ctx["known_data_from"] > 0 and ctx["mismatched_data_from"] == ctx["known_data_from"]:
            logger(
                "Warning: selected CurvePoints DataFrom looks incompatible with selected "
                f"result type prefix '{ctx['result_prefix']}'."
            )
        if curve_time_result_type_path:
            logger(f"Curve time result type: {curve_time_result_type_path}")
        else:
            logger("Curve time result type not resolved; using step/index time fallback.")

        phase_counter = 0
        for direction, phase_names in (("X", x_phase_names), ("Y", y_phase_names)):
            for phase_name in phase_names:
                phase_counter += 1
                try:
                    phase = _resolve_phase_by_name(phase_map, phase_name)
                except Exception as exc:
                    status_rows.append(
                        {
                            "Category": "PhaseResolve",
                            "Direction": direction,
                            "Phase": phase_name,
                            "Series": "",
                            "CurvePointId": "",
                            "Status": "ERROR",
                            "Message": _error_text(exc),
                        }
                    )
                    continue

                resolved_phase_name = _phase_display_name(phase)
                logger(f"[{phase_counter}/{total_phase_count}] {direction} phase: {resolved_phase_name}")
                phases_rows.append({"Direction": direction, "Phase": resolved_phase_name})
                step_list = list(phase.Steps)
                if not step_list:
                    status_rows.append(
                        {
                            "Category": "PhaseSteps",
                            "Direction": direction,
                            "Phase": resolved_phase_name,
                            "Series": "",
                            "CurvePointId": "",
                            "Status": "ERROR",
                            "Message": "Phase has no steps.",
                        }
                    )
                    logger(f"{direction} {resolved_phase_name} skipped: phase has no steps.")
                    continue

                phase_t_values, phase_dt = _resolve_phase_time_series(step_list)
                start_step = step_list[0]
                end_step = step_list[-1]
                step_count = len(step_list)
                if phase_dt is None:
                    logger(f"{resolved_phase_name}: steps={step_count}, dt=unresolved")
                else:
                    logger(f"{resolved_phase_name}: steps={step_count}, dt~{phase_dt:.6g}")

                for rec_idx, rec in enumerate(selected, start=1):
                    cp_id = rec["id"]
                    series = series_by_id.get(cp_id, rec["name"])
                    try:
                        acc_vals = _safe_float_list(
                            g_o.getcurveresultspath([rec["obj"]], start_step, end_step, acc_result_type)
                        )
                        n = len(acc_vals)
                        if n == 0:
                            raise RuntimeError("Empty acceleration series.")

                        time_vals = _read_curve_time_values(
                            g_o,
                            rec["obj"],
                            start_step,
                            end_step,
                            curve_time_result_type,
                            n,
                            step_count,
                            phase_t_values,
                            phase_dt,
                        )
                        dt = None
                        try:
                            dt = _estimate_dt(np.asarray(time_vals, dtype=float))
                        except Exception:
                            if phase_dt is not None and phase_dt > 0.0:
                                dt = phase_dt
                        snapped_dt = _snap_dt(dt, target=dt_snap_target, rel_tol=dt_snap_rel_tol)
                        if snapped_dt != dt:
                            logger(
                                f"{direction} {resolved_phase_name} | {series} "
                                f"dt snapped: {dt:.8g} -> {snapped_dt:.8g}"
                            )
                            dt = snapped_dt

                        for idx, (t_val, a_val) in enumerate(zip(time_vals, acc_vals), start=1):
                            time_rows.append(
                                {
                                    "Direction": direction,
                                    "Phase": resolved_phase_name,
                                    "CurvePointId": cp_id,
                                    "Series": series,
                                    "NodeName": rec["node_name"],
                                    "Step": idx,
                                    time_col: float(t_val),
                                    "Acceleration": float(a_val),
                                }
                            )

                        if dt is None or dt <= 0.0 or n < 2:
                            status_rows.append(
                                {
                                    "Category": "NodeRead",
                                    "Direction": direction,
                                    "Phase": resolved_phase_name,
                                    "Series": series,
                                    "CurvePointId": cp_id,
                                    "Status": "WARN",
                                    "Message": f"Insufficient time resolution for spectrum (n={n}, dt={dt}).",
                                }
                            )
                            logger(
                                f"{direction} {resolved_phase_name} | "
                                f"{series} ({rec_idx}/{len(selected)}) -> WARN: no spectrum (n={n}, dt={dt})"
                            )
                            continue

                        psa = _compute_psa_spectrum(np.asarray(acc_vals, dtype=float), dt, periods, damping)
                        for period_s, psa_val in zip(periods, psa):
                            spectrum_rows.append(
                                {
                                    "Direction": direction,
                                    "Phase": resolved_phase_name,
                                    "CurvePointId": cp_id,
                                    "Series": series,
                                    "NodeName": rec["node_name"],
                                    "Period_s": float(period_s),
                                    "PSA_g": float(psa_val),
                                }
                            )
                        status_rows.append(
                            {
                                "Category": "NodeRead",
                                "Direction": direction,
                                "Phase": resolved_phase_name,
                                "Series": series,
                                "CurvePointId": cp_id,
                                "Status": "OK",
                                "Message": f"{n} points, dt={dt:.6g}",
                            }
                        )
                        logger(
                            f"{direction} {resolved_phase_name} | "
                            f"{series} ({rec_idx}/{len(selected)}) -> OK"
                        )
                    except Exception as exc:
                        status_rows.append(
                            {
                                "Category": "NodeRead",
                                "Direction": direction,
                                "Phase": resolved_phase_name,
                                "Series": series,
                                "CurvePointId": cp_id,
                                "Status": "ERROR",
                                "Message": _error_text(exc),
                            }
                        )
                        logger(
                            f"{direction} {resolved_phase_name} | "
                            f"{series} ({rec_idx}/{len(selected)}) -> ERROR: {_error_text(exc)}"
                        )

        time_df = pd.DataFrame(time_rows)
        spectrum_long_df = pd.DataFrame(spectrum_rows)
        detail = ""
        if spectrum_long_df.empty:
            warn_or_err = pd.DataFrame(status_rows)
            if not warn_or_err.empty:
                mask = warn_or_err["Category"].isin(["PhaseSteps", "NodeRead"])
                text_list = warn_or_err[mask]["Message"].astype(str).dropna().tolist()[:3]
                detail = " | ".join(text_list).strip()
                status_rows.append(
                    {
                        "Category": "NodeSpectrum",
                        "Direction": "",
                        "Phase": "",
                        "Series": "",
                        "CurvePointId": "",
                        "Status": "WARN",
                        "Message": (
                            "No spectrum rows generated. "
                            f"{detail}" if detail else "No spectrum rows generated."
                        ),
                    }
                )
                logger(
                    "Warning: No node spectrum data generated; workbook will contain "
                    "time history and status details."
                )
            if time_df.empty:
                if detail:
                    raise RuntimeError(
                        "No node spectrum data could be read for selected phases/curve points. "
                        f"Details: {detail}"
                    )
                raise RuntimeError("No node spectrum data could be read for selected phases/curve points.")
            spectrum_long_df = pd.DataFrame(
                columns=[
                    "Direction",
                    "Phase",
                    "CurvePointId",
                    "Series",
                    "NodeName",
                    "Period_s",
                    "PSA_g",
                ]
            )

        if spectrum_long_df.empty:
            spectrum_mean_df = pd.DataFrame(
                columns=[
                    "Direction",
                    "CurvePointId",
                    "Series",
                    "NodeName",
                    "Period_s",
                    "PSA_g",
                    "PhaseCount",
                ]
            )
        else:
            spectrum_mean_df = (
                spectrum_long_df.groupby(
                    ["Direction", "CurvePointId", "Series", "NodeName", "Period_s"], as_index=False
                )
                .agg(PSA_g=("PSA_g", "mean"), PhaseCount=("Phase", "count"))
                .sort_values(["Direction", "Series", "Period_s"])
                .reset_index(drop=True)
            )

        out_path.parent.mkdir(parents=True, exist_ok=True)
        plot_dir = _ensure_plot_dir(out_path)
        chart_paths = []
        node_map_df = pd.DataFrame(node_map_rows)

        png_node_map = plot_dir / "node_selection_map.png"
        if _plot_node_selection_map(ctx["model_node_cloud"], node_map_df, png_node_map, dpi=plot_dpi):
            chart_paths.append(str(png_node_map))

        for direction in sorted(spectrum_long_df["Direction"].dropna().unique().tolist()):
            dir_time = time_df[time_df["Direction"] == direction].copy()
            dir_spec = spectrum_long_df[spectrum_long_df["Direction"] == direction].copy()
            dir_mean = spectrum_mean_df[spectrum_mean_df["Direction"] == direction].copy()

            series_list = sorted(dir_spec["Series"].dropna().unique().tolist())
            for series in series_list:
                one_time = dir_time[dir_time["Series"] == series].copy()
                one_spec = dir_spec[dir_spec["Series"] == series].copy()
                one_mean = dir_mean[dir_mean["Series"] == series].copy()

                png_time = plot_dir / f"node_time_{direction}_{safe_label(series)}.png"
                if _plot_node_timehistory(
                    direction,
                    series,
                    one_time,
                    png_time,
                    time_col,
                    dpi=plot_dpi,
                ):
                    chart_paths.append(str(png_time))

                png_spec = plot_dir / f"node_spectrum_{direction}_{safe_label(series)}.png"
                if _plot_node_spectrum_single(
                    direction,
                    series,
                    one_spec,
                    one_mean,
                    png_spec,
                    dpi=plot_dpi,
                ):
                    chart_paths.append(str(png_spec))

            png_group_phase = plot_dir / f"node_group_phase_{direction}.png"
            if _plot_node_spectrum_group_phase(
                direction,
                spectrum_long_df,
                png_group_phase,
                dpi=plot_dpi,
            ):
                chart_paths.append(str(png_group_phase))

            png_group_mean = plot_dir / f"node_group_mean_{direction}.png"
            if _plot_node_spectrum_group_mean(
                direction,
                spectrum_mean_df,
                png_group_mean,
                dpi=plot_dpi,
            ):
                chart_paths.append(str(png_group_mean))

        if save_node_timehistory_subfolders:
            written_histories = _export_node_timehistory_subfolders(
                time_df,
                out_path,
                time_col,
                logger=logger,
            )
            for path_text in written_histories:
                status_rows.append(
                    {
                        "Category": "TimeHistoryFile",
                        "Direction": "",
                        "Phase": "",
                        "Series": "",
                        "CurvePointId": "",
                        "Status": "OK",
                        "Message": path_text,
                    }
                )

        for path_text in chart_paths:
            logger(f"Chart -> {path_text}")
            status_rows.append(
                {
                    "Category": "Chart",
                    "Direction": "",
                    "Phase": "",
                    "Series": "",
                    "CurvePointId": "",
                    "Status": "OK",
                    "Message": path_text,
                }
            )

        phases_df = pd.DataFrame(phases_rows).drop_duplicates()
        selections_df = pd.DataFrame(selection_rows).drop_duplicates()
        status_df = pd.DataFrame(status_rows)
        spectrum_wide_specs = _build_node_spectrum_wide_specs(
            spectrum_long_df,
            spectrum_mean_df,
        )

        sheets = [
            ("Phases", phases_df),
            ("Selections", selections_df),
            ("NodeMap", node_map_df),
            ("NodeTimeHistoryLong", time_df),
            ("NodeSpectrumLong", spectrum_long_df),
            ("NodeSpectrumMean", spectrum_mean_df),
            ("_Status", status_df),
        ]
        for spec in spectrum_wide_specs:
            sheets.append((spec["sheet_name"], spec["frame"]))
        out_final = _write_multisheet_workbook(out_path, sheets, logger=logger)
        _add_excel_line_charts(out_final, spectrum_wide_specs, logger=logger)
        logger(f"OK -> {out_final}")
        logger(f"Charts -> {plot_dir}")
    finally:
        _safe_close_server(s_o)


def run_node_stress_strain_export(args, logger=print):
    x_phase_names = list(getattr(args, "x_phase_names", []) or [])
    y_phase_names = list(getattr(args, "y_phase_names", []) or [])
    selected_ids = list(getattr(args, "curvepoint_id", []) or [])
    out_text = str(getattr(args, "stress_strain_out", "") or getattr(args, "out", "") or "").strip()
    out_path = Path(out_text).expanduser()
    plot_dpi = int(float(getattr(args, "plot_dpi", 150) or 150))
    stress_result_type_path = str(getattr(args, "stress_result_type", "") or "").strip()
    strain_result_type_path = str(getattr(args, "strain_result_type", "") or "").strip()
    time_col = str(getattr(args, "time_col", "DynamicTime")).strip() or "DynamicTime"

    if not str(getattr(args, "password", "")).strip():
        raise RuntimeError("Password is required.")
    if not x_phase_names and not y_phase_names:
        raise RuntimeError("Select at least one X or Y phase.")
    if not out_text:
        raise RuntimeError("Stress-strain output path is required.")
    if not stress_result_type_path or not strain_result_type_path:
        raise RuntimeError("Stress and strain result types are required.")

    s_o, g_o = _open_output_server(args.host, int(args.port), args.password)
    try:
        _ensure_output_result_types(g_o)
        _ensure_active_output_project(g_o)
        stress_result_type = resolve_result_type(g_o, stress_result_type_path)
        strain_result_type = resolve_result_type(g_o, strain_result_type_path)
        curve_time_result_type, curve_time_result_type_path = _resolve_curve_time_result_type(
            g_o,
            stress_result_type_path,
            time_col,
        )

        ctx = _build_node_export_context(g_o, selected_ids, stress_result_type_path)
        selected = ctx["selected"]
        series_by_id = ctx["series_by_id"]
        phase_map = ctx["phase_map"]
        phases_rows = []
        selection_rows = list(ctx["selection_rows"])
        node_map_rows = list(ctx["node_map_rows"])
        status_rows = []
        stress_strain_rows = []

        total_phase_count = len(x_phase_names) + len(y_phase_names)
        logger(
            f"Stress-strain run started: phases={total_phase_count}, "
            f"curvepoints={len(selected)}, tau={stress_result_type_path}, "
            f"gamma={strain_result_type_path}, plot_dpi={plot_dpi}"
        )
        if ctx["known_data_from"] > 0 and ctx["mismatched_data_from"] == ctx["known_data_from"]:
            logger(
                "Warning: selected CurvePoints DataFrom looks incompatible with selected "
                f"result type prefix '{ctx['result_prefix']}'."
            )
        if curve_time_result_type_path:
            logger(f"Curve time result type: {curve_time_result_type_path}")
        else:
            logger("Curve time result type not resolved; using step/index time fallback.")

        phase_counter = 0
        for direction, phase_names in (("X", x_phase_names), ("Y", y_phase_names)):
            for phase_name in phase_names:
                phase_counter += 1
                try:
                    phase = _resolve_phase_by_name(phase_map, phase_name)
                except Exception as exc:
                    status_rows.append(
                        {
                            "Category": "PhaseResolve",
                            "Direction": direction,
                            "Phase": phase_name,
                            "Series": "",
                            "CurvePointId": "",
                            "Status": "ERROR",
                            "Message": _error_text(exc),
                        }
                    )
                    continue

                resolved_phase_name = _phase_display_name(phase)
                logger(f"[{phase_counter}/{total_phase_count}] {direction} phase: {resolved_phase_name}")
                phases_rows.append({"Direction": direction, "Phase": resolved_phase_name})
                step_list = list(phase.Steps)
                if not step_list:
                    status_rows.append(
                        {
                            "Category": "PhaseSteps",
                            "Direction": direction,
                            "Phase": resolved_phase_name,
                            "Series": "",
                            "CurvePointId": "",
                            "Status": "ERROR",
                            "Message": "Phase has no steps.",
                        }
                    )
                    logger(f"{direction} {resolved_phase_name} skipped: phase has no steps.")
                    continue

                phase_t_values, phase_dt = _resolve_phase_time_series(step_list)
                start_step = step_list[0]
                end_step = step_list[-1]
                step_count = len(step_list)

                for rec_idx, rec in enumerate(selected, start=1):
                    cp_id = rec["id"]
                    series = series_by_id.get(cp_id, rec["name"])
                    try:
                        tau_vals = _safe_float_list(
                            g_o.getcurveresultspath([rec["obj"]], start_step, end_step, stress_result_type)
                        )
                        gamma_vals = _safe_float_list(
                            g_o.getcurveresultspath([rec["obj"]], start_step, end_step, strain_result_type)
                        )
                        m = min(len(tau_vals), len(gamma_vals))
                        if m <= 0:
                            raise RuntimeError("Empty stress-strain series.")

                        time_vals = _read_curve_time_values(
                            g_o,
                            rec["obj"],
                            start_step,
                            end_step,
                            curve_time_result_type,
                            m,
                            step_count,
                            phase_t_values,
                            phase_dt,
                        )

                        for idx, (t_val, gamma_val, tau_val) in enumerate(
                            zip(time_vals[:m], gamma_vals[:m], tau_vals[:m]),
                            start=1,
                        ):
                            stress_strain_rows.append(
                                {
                                    "Direction": direction,
                                    "Phase": resolved_phase_name,
                                    "CurvePointId": cp_id,
                                    "Series": series,
                                    "NodeName": rec["node_name"],
                                    "Step": idx,
                                    time_col: float(t_val),
                                    "Gamma_xy": float(gamma_val),
                                    "Tau_xy": float(tau_val),
                                }
                            )

                        status_rows.append(
                            {
                                "Category": "StressStrainRead",
                                "Direction": direction,
                                "Phase": resolved_phase_name,
                                "Series": series,
                                "CurvePointId": cp_id,
                                "Status": "OK",
                                "Message": f"{m} points",
                            }
                        )
                        logger(
                            f"{direction} {resolved_phase_name} | "
                            f"{series} ({rec_idx}/{len(selected)}) -> OK"
                        )
                    except Exception as exc:
                        status_rows.append(
                            {
                                "Category": "StressStrainRead",
                                "Direction": direction,
                                "Phase": resolved_phase_name,
                                "Series": series,
                                "CurvePointId": cp_id,
                                "Status": "ERROR",
                                "Message": _error_text(exc),
                            }
                        )
                        logger(
                            f"{direction} {resolved_phase_name} | "
                            f"{series} ({rec_idx}/{len(selected)}) -> ERROR: {_error_text(exc)}"
                        )

        stress_strain_df = pd.DataFrame(stress_strain_rows)
        if stress_strain_df.empty:
            detail = ""
            warn_or_err = pd.DataFrame(status_rows)
            if not warn_or_err.empty:
                mask = warn_or_err["Category"].isin(["PhaseSteps", "StressStrainRead"])
                text_list = warn_or_err[mask]["Message"].astype(str).dropna().tolist()[:3]
                detail = " | ".join(text_list).strip()
            if detail:
                raise RuntimeError(
                    "No stress-strain data could be read for selected phases/curve points. "
                    f"Details: {detail}"
                )
            raise RuntimeError("No stress-strain data could be read for selected phases/curve points.")

        out_path.parent.mkdir(parents=True, exist_ok=True)
        plot_dir = _ensure_plot_dir(out_path)
        chart_paths = []
        node_map_df = pd.DataFrame(node_map_rows)

        png_node_map = plot_dir / "node_selection_map.png"
        if _plot_node_selection_map(ctx["model_node_cloud"], node_map_df, png_node_map, dpi=plot_dpi):
            chart_paths.append(str(png_node_map))

        for direction in sorted(stress_strain_df["Direction"].dropna().unique().tolist()):
            dir_df = stress_strain_df[stress_strain_df["Direction"] == direction].copy()
            for series in sorted(dir_df["Series"].dropna().unique().tolist()):
                one = dir_df[dir_df["Series"] == series].copy()
                png_loop = plot_dir / f"stress_strain_{direction}_{safe_label(series)}.png"
                if _plot_stress_strain_single(direction, series, one, png_loop, dpi=plot_dpi):
                    chart_paths.append(str(png_loop))

        for path_text in chart_paths:
            logger(f"Chart -> {path_text}")
            status_rows.append(
                {
                    "Category": "Chart",
                    "Direction": "",
                    "Phase": "",
                    "Series": "",
                    "CurvePointId": "",
                    "Status": "OK",
                    "Message": path_text,
                }
            )

        phases_df = pd.DataFrame(phases_rows).drop_duplicates()
        selections_df = pd.DataFrame(selection_rows).drop_duplicates()
        status_df = pd.DataFrame(status_rows)
        stress_strain_wide_specs = _build_stress_strain_wide_specs(stress_strain_df)

        sheets = [
            ("Phases", phases_df),
            ("Selections", selections_df),
            ("NodeMap", node_map_df),
            ("StressStrainLong", stress_strain_df),
            ("_Status", status_df),
        ]
        for spec in stress_strain_wide_specs:
            sheets.append((spec["sheet_name"], spec["frame"]))
        out_final = _write_multisheet_workbook(out_path, sheets, logger=logger)
        _add_excel_line_charts(out_final, stress_strain_wide_specs, logger=logger)
        logger(f"OK -> {out_final}")
        logger(f"Charts -> {plot_dir}")
    finally:
        _safe_close_server(s_o)

def main():
    args = parse_args()
    if args.mode == "spectrum-gui":
        run_spectrum_gui(args)
    elif args.mode == "timehistory-api":
        run_timehistory_api(args)
    elif args.mode == "curvepoints-api":
        run_curvepoints_api_export(args)
    else:
        raise RuntimeError(f"Unknown mode: {args.mode}")


if __name__ == "__main__":
    main()
